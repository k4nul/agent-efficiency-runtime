from __future__ import annotations

import io
import json
import zipfile
from pathlib import Path
from typing import Any

import openpyxl
import pytest

import aer.data.query as data_query
import aer.inspect.office as office_inspect
import aer.patch.engine as patch_engine
import aer.validation.engine as validation_engine
import aer.zip_safety as zip_safety
from aer.errors import AerError
from aer.inspect import inspect_target
from aer.patch import apply_patch
from aer.validation import validate_file
from aer.zip_safety import enforce_zip_expansion_limits


def _zip_bytes(entries: dict[str, bytes]) -> bytes:
    output = io.BytesIO()
    with zipfile.ZipFile(output, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for name, content in entries.items():
            archive.writestr(name, content)
    return output.getvalue()


def _office_bomb(path: Path, suffix: str) -> Path:
    required_parts = {
        "pptx": "ppt/presentation.xml",
        "docx": "word/document.xml",
        "xlsx": "xl/workbook.xml",
    }
    path.write_bytes(
        _zip_bytes(
            {
                "[Content_Types].xml": b"<Types/>",
                required_parts[suffix]: b"<root/>",
                "payload.bin": b"x" * 4096,
            }
        )
    )
    return path


def _unexpected_parser(*args: Any, **kwargs: Any) -> Any:
    del args, kwargs
    raise AssertionError("Office parser must not run before ZIP safety checks")


def test_common_zip_helper_checks_entry_count_and_expanded_size(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    package = _zip_bytes({"one": b"a", "two": b"b", "large": b"x" * 1024})

    monkeypatch.setattr(zip_safety, "MAX_ZIP_ENTRIES", 2)
    with pytest.raises(AerError) as entry_error:
        enforce_zip_expansion_limits(package, operation="test")
    assert entry_error.value.code == "LIMIT_EXCEEDED"
    assert entry_error.value.details == {"entries": 3, "limit": 2}

    monkeypatch.setattr(zip_safety, "MAX_ZIP_ENTRIES", 10)
    monkeypatch.setattr(zip_safety, "MAX_ZIP_UNCOMPRESSED_BYTES", 100)
    with pytest.raises(AerError) as size_error:
        enforce_zip_expansion_limits(package, operation="test")
    assert size_error.value.code == "LIMIT_EXCEEDED"
    assert size_error.value.details == {"uncompressed_bytes": 1026, "limit": 100}


def test_common_zip_helper_rejects_corrupt_input() -> None:
    with pytest.raises(AerError) as caught:
        enforce_zip_expansion_limits(b"not a zip", operation="test", target="input.docx")
    assert caught.value.code == "CORRUPT_FILE"
    assert caught.value.target == "input.docx"


def test_common_zip_helper_rejects_high_compression_ratio(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    package = _zip_bytes({"compressed.txt": b"x" * (2 * 1024 * 1024)})
    monkeypatch.setattr(zip_safety, "MAX_ZIP_COMPRESSION_RATIO", 10)

    with pytest.raises(AerError) as caught:
        enforce_zip_expansion_limits(package, operation="test")

    assert caught.value.code == "LIMIT_EXCEEDED"
    assert caught.value.target == "compressed.txt"


@pytest.mark.parametrize(
    ("suffix", "parser_name"),
    [("pptx", "Presentation"), ("docx", "Document"), ("xlsx", "load_workbook")],
)
def test_validation_rejects_office_bomb_before_parser(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    suffix: str,
    parser_name: str,
) -> None:
    target = _office_bomb(tmp_path / f"bomb.{suffix}", suffix)
    monkeypatch.setattr(zip_safety, "MAX_ZIP_UNCOMPRESSED_BYTES", 100)
    monkeypatch.setattr(validation_engine, parser_name, _unexpected_parser)

    with pytest.raises(AerError) as caught:
        validate_file(target)

    assert caught.value.code == "LIMIT_EXCEEDED"
    assert caught.value.target == str(target)


@pytest.mark.parametrize(
    ("suffix", "parser_name"),
    [("pptx", "Presentation"), ("docx", "Document"), ("xlsx", "load_workbook")],
)
def test_inspect_rejects_office_bomb_before_parser(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    suffix: str,
    parser_name: str,
) -> None:
    target = _office_bomb(tmp_path / f"bomb.{suffix}", suffix)
    monkeypatch.setattr(zip_safety, "MAX_ZIP_UNCOMPRESSED_BYTES", 100)
    monkeypatch.setattr(office_inspect, parser_name, _unexpected_parser)

    with pytest.raises(AerError) as caught:
        inspect_target(target)

    assert caught.value.code == "LIMIT_EXCEEDED"
    assert caught.value.operation == "inspect"


def test_data_query_rejects_xlsx_bomb_before_parser(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    target = _office_bomb(tmp_path / "bomb.xlsx", "xlsx")
    monkeypatch.setattr(zip_safety, "MAX_ZIP_UNCOMPRESSED_BYTES", 100)
    monkeypatch.setattr(openpyxl, "load_workbook", _unexpected_parser)

    with pytest.raises(AerError) as caught:
        data_query.query_data(target)

    assert caught.value.code == "LIMIT_EXCEEDED"
    assert caught.value.operation == "data.query"


@pytest.mark.parametrize(
    ("suffix", "parser_name", "operation"),
    [
        ("pptx", "Presentation", {"op": "pptx.set_text", "target": "slide:1/shape:1"}),
        ("docx", "Document", {"op": "docx.replace_text", "old": "a", "value": "b"}),
        ("xlsx", "load_workbook", {"op": "xlsx.set_cell", "target": "Sheet!A1", "value": 1}),
    ],
)
def test_patch_rejects_office_bomb_before_parser_and_preserves_source(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    suffix: str,
    parser_name: str,
    operation: dict[str, object],
) -> None:
    target = _office_bomb(tmp_path / f"bomb.{suffix}", suffix)
    before = target.read_bytes()
    spec = tmp_path / f"{suffix}-patch.json"
    spec.write_text(json.dumps({"version": 1, "operations": [operation]}), encoding="utf-8")
    monkeypatch.setattr(zip_safety, "MAX_ZIP_UNCOMPRESSED_BYTES", 100)
    monkeypatch.setattr(patch_engine, parser_name, _unexpected_parser)

    with pytest.raises(AerError) as caught:
        apply_patch(target, spec)

    assert caught.value.code == "LIMIT_EXCEEDED"
    assert caught.value.operation == "artifact.patch"
    assert target.read_bytes() == before
