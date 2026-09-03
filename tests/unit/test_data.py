from __future__ import annotations

import csv
import json
from dataclasses import dataclass
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

import aer.data.query as data_query_module
from aer.data import AggregateSpec, data_response, parse_aggregate, parse_filter, query_data
from aer.errors import AerError


@dataclass
class _Stored:
    ref: str


class _MemoryStore:
    def __init__(self) -> None:
        self.data = b""
        self.filename: str | None = None

    def put_bytes(
        self,
        data: bytes,
        *,
        filename: str | None = None,
        mime_type: str | None = None,
        source: dict[str, object] | None = None,
    ) -> _Stored:
        self.data = data
        self.filename = filename
        return _Stored("aer://sha256/" + "b" * 64)


@pytest.fixture
def people_csv(tmp_path: Path) -> Path:
    path = tmp_path / "people.csv"
    path.write_text(
        "id,status,total,name,note,team\n"
        "1,pending,43000,Alpha,hello world,A\n"
        "2,paid,12000,Beta,,A\n"
        "3,pending,31000,Alpine,final note,B\n"
        "4,cancelled,500,Gamma,hello,B\n",
        encoding="utf-8",
    )
    return path


@pytest.mark.parametrize(
    ("expression", "expected"),
    [
        ("status == pending", [1, 3]),
        ("status != pending", [2, 4]),
        ("total > 12000", [1, 3]),
        ("total >= 12000", [1, 2, 3]),
        ("total < 12000", [4]),
        ("total <= 12000", [2, 4]),
        ("note contains hello", [1, 4]),
        ("name starts_with Al", [1, 3]),
        ("name ends_with ta", [2]),
        ("note is_null", [2]),
        ("note not_null", [1, 3, 4]),
        ("status in pending,paid", [1, 2, 3]),
        ('status in ["paid", "cancelled"]', [2, 4]),
    ],
)
def test_all_filter_operators(people_csv: Path, expression: str, expected: list[int]) -> None:
    result = query_data(people_csv, where=[expression], select="id")
    assert [row["id"] for row in result.preview] == expected


def test_filter_parser_never_executes_python(tmp_path: Path, people_csv: Path) -> None:
    marker = tmp_path / "owned"
    expression = f"status == __import__('pathlib').Path('{marker}').touch()"
    result = query_data(people_csv, where=[expression])
    assert result.matched_rows == 0
    assert not marker.exists()
    with pytest.raises(AerError, match="Invalid filter"):
        parse_filter("status or __import__('os').system('id')")


def test_select_rename_sort_offset_limit(people_csv: Path) -> None:
    result = query_data(
        people_csv,
        where=["status == pending"],
        select="id,total",
        rename={"total": "amount"},
        sort="amount",
        descending=True,
        offset=1,
        limit=1,
    )
    assert result.matched_rows == 2
    assert result.columns == ["id", "amount"]
    assert result.preview == [{"id": 3, "amount": 31000}]


def test_csv_decimals_are_json_serializable_and_nulls_do_not_compare(tmp_path: Path) -> None:
    source = tmp_path / "decimal.csv"
    source.write_text("id,value\n1,1.25\n2,\n")
    result = query_data(source, where="value < 2")
    assert result.preview == [{"id": 1, "value": 1.25}]
    assert json.loads(json.dumps(result.to_dict()))["preview"][0]["value"] == 1.25


def test_unique_and_duplicate_detection(tmp_path: Path) -> None:
    source = tmp_path / "duplicates.json"
    source.write_text(
        json.dumps(
            [
                {"id": 1, "name": "same"},
                {"id": 2, "name": "same"},
                {"id": 3, "name": "other"},
            ]
        )
    )
    duplicates = query_data(source, duplicate_columns="name")
    assert duplicates.duplicate_rows == 2
    assert [row["id"] for row in duplicates.preview] == [1, 2]
    unique = query_data(source, unique="name")
    assert [row["id"] for row in unique.preview] == [1, 3]


def test_grouped_aggregates(tmp_path: Path) -> None:
    source = tmp_path / "values.json"
    source.write_text(
        json.dumps(
            [
                {"team": "A", "value": 10},
                {"team": "A", "value": 20},
                {"team": "A", "value": None},
                {"team": "B", "value": 5},
            ]
        )
    )
    result = query_data(
        source,
        group_by="team",
        aggregates=[
            "count",
            "sum:value",
            "average:value",
            "min:value",
            "max:value",
            "null_count:value",
        ],
        sort="team",
    )
    assert result.preview == [
        {
            "team": "A",
            "count": 3,
            "sum_value": 30,
            "average_value": 15,
            "min_value": 10,
            "max_value": 20,
            "null_count_value": 1,
        },
        {
            "team": "B",
            "count": 1,
            "sum_value": 5,
            "average_value": 5,
            "min_value": 5,
            "max_value": 5,
            "null_count_value": 0,
        },
    ]


def test_ten_thousand_rows_return_twenty_row_preview_and_ref(tmp_path: Path) -> None:
    source = tmp_path / "large.csv"
    with source.open("w", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(["id", "status"])
        writer.writerows((index, "pending") for index in range(10_000))
    store = _MemoryStore()
    result = query_data(source, where=["status == pending"], store=store)

    assert result.matched_rows == 10_000
    assert result.result_rows == 10_000
    assert len(result.preview) == 20
    assert result.result_ref == "aer://sha256/" + "b" * 64
    assert len(store.data.splitlines()) == 10_000
    assert len(json.dumps(result.to_dict()).encode()) < 16 * 1024


@pytest.mark.parametrize("suffix", [".tsv", ".json", ".jsonl"])
def test_supported_text_loaders(tmp_path: Path, suffix: str) -> None:
    source = tmp_path / f"data{suffix}"
    if suffix == ".tsv":
        source.write_text("id\tvalue\n1\t10\n2\t20\n")
    elif suffix == ".json":
        source.write_text('[{"id":1,"value":10},{"id":2,"value":20}]')
    else:
        source.write_text('{"id":1,"value":10}\n{"id":2,"value":20}\n')
    result = query_data(source, where=["value >= 20"])
    assert result.preview == [{"id": 2, "value": 20}]


@pytest.mark.parametrize("suffix", [".csv", ".tsv", ".json", ".jsonl"])
def test_supported_text_outputs_are_parseable(
    tmp_path: Path, people_csv: Path, suffix: str
) -> None:
    output = tmp_path / f"result{suffix}"
    result = query_data(people_csv, select="id,name", limit=2, output=output)
    assert result.result_rows == 2
    if suffix == ".csv":
        assert list(csv.DictReader(output.open())) == [
            {"id": "1", "name": "Alpha"},
            {"id": "2", "name": "Beta"},
        ]
    elif suffix == ".tsv":
        assert list(csv.DictReader(output.open(), delimiter="\t"))[1]["name"] == "Beta"
    elif suffix == ".json":
        assert json.loads(output.read_text())[0] == {"id": 1, "name": "Alpha"}
    else:
        assert [json.loads(line) for line in output.read_text().splitlines()][1] == {
            "id": 2,
            "name": "Beta",
        }


def test_xlsx_loader_preserves_formula_string(tmp_path: Path) -> None:
    source = tmp_path / "input.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Raw"
    worksheet.append(["id", "amount", "formula"])
    worksheet.append([1, 50, "=B2*2"])
    worksheet.append([2, 100, "=B3*2"])
    workbook.save(source)

    result = query_data(source, sheet="Raw", where=["amount >= 100"])
    assert result.preview == [{"id": 2, "amount": 100, "formula": "=B3*2"}]


def test_outputs_are_reopenable_and_stored(tmp_path: Path, people_csv: Path) -> None:
    output = tmp_path / "result.xlsx"
    store = _MemoryStore()
    result = query_data(
        people_csv,
        select="id,total",
        limit=2,
        output=output,
        store=store,
    )
    workbook = load_workbook(output, read_only=True)
    try:
        assert list(workbook.active.values) == [("id", "total"), (1, 43000), (2, 12000)]
    finally:
        workbook.close()
    assert result.output == str(output)
    assert result.bytes_written == output.stat().st_size
    assert store.data == output.read_bytes()


def test_data_response_is_compact(people_csv: Path) -> None:
    result = query_data(people_csv)
    response = data_response(result)
    assert response["ok"] is True
    assert response["operation"] == "data.query"
    assert "rows" not in response["result"]


def test_loader_and_query_errors_are_actionable(tmp_path: Path, people_csv: Path) -> None:
    invalid_json = tmp_path / "bad.json"
    invalid_json.write_text("not-json")
    with pytest.raises(AerError) as corrupt:
        query_data(invalid_json)
    assert corrupt.value.code == "CORRUPT_FILE"

    unsupported = tmp_path / "data.yaml"
    unsupported.write_text("items: []")
    with pytest.raises(AerError) as unsupported_error:
        query_data(unsupported)
    assert unsupported_error.value.code == "UNSUPPORTED_FORMAT"

    with pytest.raises(AerError, match="group_by requires"):
        query_data(people_csv, group_by="team")
    with pytest.raises(AerError, match="non-negative"):
        query_data(people_csv, limit=-1)


def test_duplicate_headers_and_missing_xlsx_sheet_are_rejected(tmp_path: Path) -> None:
    duplicate_headers = tmp_path / "duplicate.csv"
    duplicate_headers.write_text("id,id\n1,2\n")
    with pytest.raises(AerError, match="duplicate column"):
        query_data(duplicate_headers)

    workbook_path = tmp_path / "book.xlsx"
    workbook = Workbook()
    workbook.active.append(["id"])
    workbook.save(workbook_path)
    with pytest.raises(AerError) as missing_sheet:
        query_data(workbook_path, sheet="Missing")
    assert missing_sheet.value.code == "NOT_FOUND"


def test_aggregate_parser_validates_shape() -> None:
    assert parse_aggregate("sum(value) as total") == AggregateSpec("sum", "value", "total")
    assert parse_aggregate("avg:value") == AggregateSpec("average", "value")
    with pytest.raises(AerError, match="requires a column"):
        parse_aggregate("sum")
    with pytest.raises(AerError, match="does not accept"):
        parse_aggregate("count:value")


def test_invalid_columns_and_non_numeric_aggregates(people_csv: Path) -> None:
    with pytest.raises(AerError, match="do not exist"):
        query_data(people_csv, where=["missing == value"])
    with pytest.raises(AerError, match="requires numeric"):
        query_data(people_csv, aggregates=[AggregateSpec("sum", "status")])
    with pytest.raises(AerError, match="duplicate output columns"):
        query_data(
            people_csv,
            aggregates=["sum:total as amount", "max:total as amount"],
        )


def test_input_output_conflict_is_rejected(people_csv: Path) -> None:
    with pytest.raises(AerError) as caught:
        query_data(people_csv, output=people_csv)
    assert caught.value.code == "CONFLICT"


def test_xlsx_output_preserves_formula_like_csv_values_as_text(tmp_path: Path) -> None:
    source = tmp_path / "formula.csv"
    source.write_text('id,payload\n1,=HYPERLINK("https://example.test")\n', encoding="utf-8")
    output = tmp_path / "formula.xlsx"

    query_data(source, output=output)

    workbook = load_workbook(output, data_only=False)
    try:
        cell = workbook.active["B2"]
        assert cell.value == '=HYPERLINK("https://example.test")'
        assert cell.data_type == "s"
        assert cell.quotePrefix is True
    finally:
        workbook.close()


@pytest.mark.parametrize(
    ("suffix", "content"),
    [
        (".json", '[{"a":1},{"b":2},{"c":3}]'),
        (".jsonl", '{"a":1}\n{"b":2}\n{"c":3}\n'),
    ],
)
def test_sparse_json_tables_are_rejected_before_rectangular_expansion(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
    suffix: str,
    content: str,
) -> None:
    source = tmp_path / f"sparse{suffix}"
    source.write_text(content, encoding="utf-8")
    monkeypatch.setattr(data_query_module, "MAX_TABULAR_CELLS", 8)

    with pytest.raises(AerError) as limit:
        query_data(source)

    assert limit.value.code == "LIMIT_EXCEEDED"
    assert limit.value.details == {
        "rows": 3,
        "columns": 3,
        "cells": 9,
        "limit": 8,
    }
