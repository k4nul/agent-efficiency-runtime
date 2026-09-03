from __future__ import annotations

import json
import subprocess
import zipfile
from pathlib import Path
from types import SimpleNamespace

import pytest
import yaml
from openpyxl import Workbook
from PIL import Image
from pptx import Presentation
from pypdf import PageObject, PdfWriter

import aer.pdf.safety as pdf_safety
import aer.validation.engine as validation_engine
from aer.artifacts import build_artifact
from aer.errors import AerError
from aer.validation import validate_file


def _yaml(path: Path, value: dict[str, object]) -> Path:
    path.write_text(yaml.safe_dump(value, sort_keys=False), encoding="utf-8")
    return path


def _blank_pdf(path: Path, pages: int = 1) -> Path:
    writer = PdfWriter()
    for _ in range(pages):
        writer.add_blank_page(width=612, height=792)
    with path.open("wb") as handle:
        writer.write(handle)
    return path


def test_generated_office_artifacts_pass_structural_validation(tmp_path: Path) -> None:
    cases = [
        (
            "presentation",
            "pptx",
            {"content": [{"id": "cover", "layout": "title", "title": "Valid"}]},
            "slide_count",
        ),
        (
            "document",
            "docx",
            {"content": [{"id": "body", "type": "paragraph", "text": "Valid"}]},
            "paragraph_count",
        ),
        (
            "workbook",
            "xlsx",
            {"sheets": [{"id": "data", "name": "Data", "columns": ["id"], "rows": [[1]]}]},
            "sheet_count",
        ),
    ]
    for kind, suffix, body, expected_check in cases:
        spec = _yaml(tmp_path / f"{kind}.yaml", {"version": 1, "kind": kind, **body})
        output = tmp_path / f"artifact.{suffix}"
        build_artifact(spec, output)
        result = validate_file(output)
        assert result["valid"] is True
        assert result["checks"][expected_check] == 1
        assert result["automatic_checks_only"] is True
        assert result["human_visual_review_required"] is True


def test_empty_presentation_is_not_reported_valid(tmp_path: Path) -> None:
    target = tmp_path / "empty.pptx"
    Presentation().save(target)

    with pytest.raises(AerError) as captured:
        validate_file(target)

    assert captured.value.code == "VALIDATION_FAILED"
    assert captured.value.details["errors"] == [
        {"code": "VALIDATION_FAILED", "message": "Presentation has no slides."}
    ]


def test_xlsx_validation_rejects_misordered_formula_parentheses(tmp_path: Path) -> None:
    target = tmp_path / "formula.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "=)("
    sheet["A2"] = '=")"&"("'
    workbook.create_sheet("Q1)")["A1"] = 1
    sheet["A3"] = "='Q1)'!A1"
    workbook.save(target)
    workbook.close()

    with pytest.raises(AerError) as captured:
        validate_file(target)

    assert captured.value.code == "VALIDATION_FAILED"
    formula_error = next(
        error
        for error in captured.value.details["errors"]
        if error["message"] == "Basic formula syntax check failed."
    )
    assert formula_error["details"]["cells"] == ["Sheet!A1"]


def test_xlsx_validation_does_not_expand_sparse_rectangles_and_limits_real_cells(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    sparse = tmp_path / "sparse.xlsx"
    workbook = Workbook()
    workbook.active["XFD1048576"] = 1
    workbook.save(sparse)
    workbook.close()

    sparse_result = validate_file(sparse)

    assert sparse_result["valid"] is True
    assert sparse_result["checks"]["materialized_cells"] == 1

    bounded = tmp_path / "bounded.xlsx"
    workbook = Workbook()
    workbook.active["A1"] = 1
    workbook.active["A2"] = 2
    workbook.save(bounded)
    workbook.close()
    monkeypatch.setattr(validation_engine, "MAX_TABULAR_CELLS", 1)

    with pytest.raises(AerError) as captured:
        validate_file(bounded)

    assert captured.value.code == "LIMIT_EXCEEDED"
    assert captured.value.details == {"cells": 2, "limit": 1}


def test_generated_svg_chart_passes_structural_validation(tmp_path: Path) -> None:
    data = tmp_path / "chart.csv"
    data.write_text("label,value\nA,1\nB,2\n", encoding="utf-8")
    spec = _yaml(
        tmp_path / "chart.yaml",
        {
            "version": 1,
            "kind": "chart",
            "type": "bar",
            "source": "chart.csv",
            "x": "label",
            "y": "value",
        },
    )
    target = tmp_path / "chart.svg"
    build_artifact(spec, target)

    result = validate_file(target)

    assert result["valid"] is True
    assert result["checks"]["root"] == "svg"
    assert result["checks"]["element_count"] > 1
    assert result["human_visual_review_required"] is True


def test_malformed_svg_is_rejected(tmp_path: Path) -> None:
    target = tmp_path / "broken.svg"
    target.write_text("<svg><g>", encoding="utf-8")

    with pytest.raises(AerError) as captured:
        validate_file(target)

    assert captured.value.code == "CORRUPT_FILE"


@pytest.mark.parametrize(
    ("limit_name", "limit"),
    [("MAX_SVG_ELEMENTS", 2), ("MAX_SVG_DEPTH", 2)],
)
def test_svg_structure_limits_are_enforced_before_full_materialization(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path, limit_name: str, limit: int
) -> None:
    target = tmp_path / "bounded.svg"
    target.write_text('<svg xmlns="http://www.w3.org/2000/svg"><g><path d="M0 0"/></g></svg>')
    monkeypatch.setattr(validation_engine, limit_name, limit)

    with pytest.raises(AerError) as captured:
        validate_file(target)

    assert captured.value.code == "LIMIT_EXCEEDED"


def test_manifest_hash_mismatch_is_validation_failure(tmp_path: Path) -> None:
    spec = _yaml(
        tmp_path / "document.yaml",
        {
            "version": 1,
            "kind": "document",
            "content": [{"id": "body", "type": "paragraph", "text": "Valid"}],
        },
    )
    target = tmp_path / "document.docx"
    build_artifact(spec, target)
    manifest_path = target.with_name(target.name + ".aer.json")
    manifest = json.loads(manifest_path.read_text())
    manifest["artifact_sha256"] = "0" * 64
    manifest_path.write_text(json.dumps(manifest))

    with pytest.raises(AerError) as caught:
        validate_file(target)
    assert caught.value.code == "VALIDATION_FAILED"
    assert caught.value.details["errors"][0]["code"] == "HASH_MISMATCH"


def test_shape_outside_slide_is_detected(tmp_path: Path) -> None:
    spec = _yaml(
        tmp_path / "outside.yaml",
        {
            "version": 1,
            "kind": "presentation",
            "content": [
                {
                    "id": "bad-layout",
                    "layout": "bullets",
                    "title": "Outside",
                    "items": ["item"],
                    "position": {"x": 13, "y": 2, "w": 2, "h": 2},
                }
            ],
        },
    )
    target = tmp_path / "outside.pptx"
    build_artifact(spec, target)
    with pytest.raises(AerError) as caught:
        validate_file(target)
    assert caught.value.code == "VALIDATION_FAILED"
    assert any(
        error["code"] == "VALIDATION_FAILED" and "outside" in error["message"].casefold()
        for error in caught.value.details["errors"]
    )


@pytest.mark.parametrize("suffix", ["pptx", "docx", "xlsx"])
def test_corrupt_ooxml_is_never_reported_valid(tmp_path: Path, suffix: str) -> None:
    target = tmp_path / f"corrupt.{suffix}"
    target.write_bytes(b"not a ZIP package")
    with pytest.raises(AerError) as caught:
        validate_file(target)
    assert caught.value.code == "CORRUPT_FILE"


def test_ooxml_missing_required_parts_is_rejected(tmp_path: Path) -> None:
    target = tmp_path / "missing.pptx"
    with zipfile.ZipFile(target, "w") as archive:
        archive.writestr("[Content_Types].xml", "<Types/>")
    with pytest.raises(AerError) as caught:
        validate_file(target)
    assert caught.value.code == "CORRUPT_FILE"


def test_corrupt_pdf_is_rejected_and_blank_pdf_warning_is_strict(tmp_path: Path) -> None:
    corrupt = tmp_path / "corrupt.pdf"
    corrupt.write_bytes(b"%PDF-1.7\nnot a real PDF")
    with pytest.raises(AerError) as caught:
        validate_file(corrupt)
    assert caught.value.code == "CORRUPT_FILE"

    blank = _blank_pdf(tmp_path / "blank.pdf")
    result = validate_file(blank)
    assert result["valid"] is True
    assert result["checks"]["page_count"] == 1
    assert result["warnings"][0]["code"] == "EMPTY_PAGE_CANDIDATE"
    with pytest.raises(AerError) as strict:
        validate_file(blank, strict=True)
    assert strict.value.code == "VALIDATION_FAILED"
    assert strict.value.details["errors"] == []


def test_pdf_validation_enforces_input_bytes_before_parsing(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    target = _blank_pdf(tmp_path / "bounded.pdf")
    monkeypatch.setattr(pdf_safety, "MAX_PDF_INPUT_BYTES", target.stat().st_size - 1)

    with pytest.raises(AerError) as caught:
        validate_file(target)

    assert caught.value.code == "LIMIT_EXCEEDED"
    assert caught.value.operation == "artifact.validate"


def test_pdf_validation_limits_text_pages_and_never_decodes_page_images(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    target = _blank_pdf(tmp_path / "bounded-pages.pdf", pages=2)
    monkeypatch.setattr(validation_engine, "MAX_PDF_TEXT_VALIDATION_PAGES", 1)

    def reject_image_decode(_page: PageObject) -> object:
        raise AssertionError("validation must not decode page images")

    monkeypatch.setattr(PageObject, "images", property(reject_image_decode))

    result = validate_file(target)

    assert result["checks"]["text_pages_checked"] == 1
    assert [warning["code"] for warning in result["warnings"]] == [
        "EMPTY_PAGE_CANDIDATE",
        "PDF_TEXT_CHECK_LIMITED",
    ]


def test_pdf_render_requires_rasterizer(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    target = _blank_pdf(tmp_path / "blank.pdf")
    monkeypatch.setattr("aer.validation.engine.shutil.which", lambda _name: None)

    with pytest.raises(AerError) as captured:
        validate_file(target, render=True)

    assert captured.value.code == "DEPENDENCY_MISSING"
    assert captured.value.details == {
        "dependency": "pdftoppm",
        "capability": "pdf.render_validate",
    }


def test_pdf_render_stores_bounded_preview_refs(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    target = _blank_pdf(tmp_path / "two-pages.pdf", pages=2)
    monkeypatch.setenv("AER_HOME", str(tmp_path / "aer-home"))
    monkeypatch.setattr("aer.validation.engine.shutil.which", lambda _name: "/usr/bin/pdftoppm")

    def fake_run(argv: list[str], **_kwargs: object) -> SimpleNamespace:
        prefix = Path(argv[-1])
        for page in (1, 2):
            Image.new("RGB", (612, 792), "white").save(prefix.with_name(f"page-{page}.png"))
        return SimpleNamespace(returncode=0)

    monkeypatch.setattr("aer.validation.engine.subprocess.run", fake_run)

    result = validate_file(target, render=True)

    render = result["checks"]["render"]
    assert render["rasterized"] is True
    assert render["pages_checked"] == 2
    assert len(render["preview_refs"]) == 2
    assert all(ref.startswith("aer://sha256/") for ref in render["preview_refs"])
    assert render["human_visual_review_required"] is True


def test_office_render_reports_missing_libreoffice(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    spec = _yaml(
        tmp_path / "document.yaml",
        {
            "version": 1,
            "kind": "document",
            "content": [{"id": "body", "type": "paragraph", "text": "Valid"}],
        },
    )
    target = tmp_path / "document.docx"
    build_artifact(spec, target)
    monkeypatch.setattr("aer.validation.engine.shutil.which", lambda _name: None)

    with pytest.raises(AerError) as captured:
        validate_file(target, render=True)

    assert captured.value.code == "DEPENDENCY_MISSING"
    assert captured.value.details["capability"] == "office.render_validate"


def test_office_render_uses_isolated_safe_mode_and_reopens_pdf(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    spec = _yaml(
        tmp_path / "document.yaml",
        {
            "version": 1,
            "kind": "document",
            "content": [{"id": "body", "type": "paragraph", "text": "Valid"}],
        },
    )
    target = tmp_path / "document.docx"
    build_artifact(spec, target)
    monkeypatch.setattr(
        "aer.validation.engine.shutil.which",
        lambda name: "/usr/bin/libreoffice" if name == "libreoffice" else None,
    )

    def fake_run(argv: list[str], **kwargs: object) -> SimpleNamespace:
        assert kwargs["shell"] is False
        assert "--safe-mode" in argv
        assert any(value.startswith("-env:UserInstallation=file:") for value in argv)
        output_dir = Path(argv[argv.index("--outdir") + 1])
        _blank_pdf(output_dir / "document.pdf", pages=2)
        return SimpleNamespace(returncode=0)

    monkeypatch.setattr("aer.validation.engine.subprocess.run", fake_run)

    result = validate_file(target, render=True)

    assert result["checks"]["render"] == {
        "rendered": True,
        "page_count": 2,
        "human_visual_review_required": True,
        "raster_preview": {"available": False, "dependency": "pdftoppm"},
    }


def test_office_render_timeout_is_mapped_to_compact_error(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    spec = _yaml(
        tmp_path / "timeout-document.yaml",
        {
            "version": 1,
            "kind": "document",
            "content": [{"id": "body", "type": "paragraph", "text": "Valid"}],
        },
    )
    target = tmp_path / "timeout-document.docx"
    build_artifact(spec, target)
    monkeypatch.setattr(
        "aer.validation.engine.shutil.which",
        lambda name: "/usr/bin/libreoffice" if name == "libreoffice" else None,
    )

    def time_out(argv: list[str], **kwargs: object) -> SimpleNamespace:
        raise subprocess.TimeoutExpired(argv, float(kwargs["timeout"]))

    monkeypatch.setattr("aer.validation.engine.subprocess.run", time_out)

    with pytest.raises(AerError) as captured:
        validate_file(target, render=True)

    assert captured.value.code == "COMMAND_TIMEOUT"
    assert captured.value.operation == "artifact.validate"
    assert captured.value.target == str(target.resolve())
    assert captured.value.details == {"timeout_seconds": 120}


def test_malformed_archive_relationship_is_rejected(tmp_path: Path) -> None:
    target = tmp_path / "relationships.zip"
    with zipfile.ZipFile(target, "w") as archive:
        archive.writestr("_rels/.rels", "<not-closed")
        archive.writestr("payload.txt", "safe")
    with pytest.raises(AerError) as caught:
        validate_file(target)
    assert caught.value.code == "VALIDATION_FAILED"
    assert caught.value.details["errors"][0]["target"] == "_rels/.rels"


def test_empty_and_unsupported_files_have_specific_errors(tmp_path: Path) -> None:
    empty = tmp_path / "empty.pdf"
    empty.touch()
    with pytest.raises(AerError) as empty_error:
        validate_file(empty)
    assert empty_error.value.code == "VALIDATION_FAILED"

    unsupported = tmp_path / "data.bin"
    unsupported.write_bytes(b"binary")
    with pytest.raises(AerError) as unsupported_error:
        validate_file(unsupported)
    assert unsupported_error.value.code == "UNSUPPORTED_FORMAT"


def test_image_content_must_match_extension(tmp_path: Path) -> None:
    mismatched = tmp_path / "actually-png.jpg"
    Image.new("RGB", (8, 8), "white").save(mismatched, format="PNG")

    with pytest.raises(AerError) as captured:
        validate_file(mismatched)

    assert captured.value.code == "VALIDATION_FAILED"
    errors = captured.value.details["errors"]
    assert errors[0]["details"] == {"extension": ".jpg", "detected_format": "PNG"}


def test_raster_image_validation_requires_human_visual_review(tmp_path: Path) -> None:
    target = tmp_path / "image.png"
    Image.new("RGB", (8, 8), "white").save(target)

    result = validate_file(target)

    assert result["valid"] is True
    assert result["automatic_checks_only"] is True
    assert result["human_visual_review_required"] is True
