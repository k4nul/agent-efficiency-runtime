from __future__ import annotations

import json
import os
import zipfile
from pathlib import Path

import pytest
from docx import Document
from openpyxl import Workbook, load_workbook
from PIL import Image
from pypdf import PdfReader, PdfWriter

import aer.conversion.ops as conversion_ops
import aer.image.ops as image_ops
from aer.archive import create_archive, list_archive, verify_archive
from aer.conversion import convert_file
from aer.errors import AerError
from aer.hashing import sha256_file
from aer.image import batch_images, crop_image, fit_image, inspect_image, resize_image
from aer.pdf import extract_pdf, inspect_pdf, merge_pdfs, parse_pages, split_pdf


def _image(path: Path, size: tuple[int, int] = (120, 80), color: str = "#2769C1") -> Path:
    Image.new("RGB", size, color).save(path)
    return path


def _pdf(path: Path, pages: int, *, title: str = "AER PDF") -> Path:
    writer = PdfWriter()
    writer.add_metadata({"/Title": title})
    for index in range(pages):
        writer.add_blank_page(width=600 + index, height=800 + index)
    with path.open("wb") as handle:
        writer.write(handle)
    return path


def test_image_resize_crop_fit_and_inspect_reopen(tmp_path: Path) -> None:
    source = _image(tmp_path / "source.png")
    resized = tmp_path / "resized.webp"
    cropped = tmp_path / "cropped.png"
    covered = tmp_path / "cover.png"
    contained = tmp_path / "contain.png"

    assert resize_image(source, resized, width=60)["height"] == 40
    assert crop_image(source, cropped, x=10, y=10, width=50, height=30)["width"] == 50
    assert fit_image(source, covered, ratio="1:1", mode="cover", width=64)["height"] == 64
    assert fit_image(source, contained, ratio="4:5", mode="contain", width=80)["height"] == 100

    for path, size in (
        (resized, (60, 40)),
        (cropped, (50, 30)),
        (covered, (64, 64)),
        (contained, (80, 100)),
    ):
        with Image.open(path) as image:
            image.verify()
        inspected = inspect_image(path)
        assert (inspected["width"], inspected["height"]) == size
        assert (
            inspected["format"]
            == {
                ".png": "PNG",
                ".webp": "WEBP",
            }[path.suffix]
        )

    with pytest.raises(AerError) as conflict:
        resize_image(source, resized, width=20)
    assert conflict.value.code == "CONFLICT"


def test_image_exif_orientation_and_pixel_limit(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    oriented = tmp_path / "oriented.jpg"
    image = Image.new("RGB", (40, 20), "red")
    exif = image.getexif()
    exif[274] = 6
    image.save(oriented, exif=exif)
    info = inspect_image(oriented)
    assert (info["width"], info["height"]) == (20, 40)
    assert info["format"] == "JPEG"

    oversized = _image(tmp_path / "oversized.png", (11, 10))
    monkeypatch.setattr(image_ops, "MAX_IMAGE_PIXELS", 100)
    with pytest.raises(AerError) as limit:
        inspect_image(oversized)
    assert limit.value.code == "LIMIT_EXCEEDED"

    with pytest.raises(AerError) as invalid_width:
        fit_image(oriented, tmp_path / "invalid.png", ratio="1:1", width=-1)
    assert invalid_width.value.code == "INVALID_ARGUMENT"

    with pytest.raises(AerError) as oversized_output:
        resize_image(oriented, tmp_path / "huge.png", width=100_000_000)
    assert oversized_output.value.code == "LIMIT_EXCEEDED"
    assert not (tmp_path / "huge.png").exists()


def test_image_batch_has_bounded_manifest_and_reopenable_outputs(tmp_path: Path) -> None:
    source_dir = tmp_path / "images"
    source_dir.mkdir()
    _image(source_dir / "a.png", (100, 50), "red")
    _image(source_dir / "b.png", (60, 60), "blue")
    output_dir = tmp_path / "batch"
    result = batch_images(str(source_dir / "*.png"), output_dir, width=40)

    assert result["count"] == 2
    manifest = json.loads((output_dir / "manifest.json").read_text())
    assert len(manifest["files"]) == 2
    with Image.open(output_dir / "a.png") as first:
        assert first.size == (40, 20)


def test_image_batch_does_not_overwrite_existing_manifest_by_default(tmp_path: Path) -> None:
    source_dir = tmp_path / "images"
    source_dir.mkdir()
    _image(source_dir / "new.png")
    output_dir = tmp_path / "batch"
    output_dir.mkdir()
    manifest = output_dir / "manifest.json"
    manifest.write_text("DO NOT OVERWRITE", encoding="utf-8")

    with pytest.raises(AerError) as conflict:
        batch_images(str(source_dir / "*.png"), output_dir, width=40)

    assert conflict.value.code == "CONFLICT"
    assert manifest.read_text(encoding="utf-8") == "DO NOT OVERWRITE"
    assert not (output_dir / "new.png").exists()


def test_image_batch_limits_and_failure_leave_no_partial_outputs(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    source_dir = tmp_path / "images"
    source_dir.mkdir()
    _image(source_dir / "a.png")
    _image(source_dir / "b.png")

    monkeypatch.setattr(image_ops, "MAX_IMAGE_BATCH_FILES", 1)
    with pytest.raises(AerError) as too_many:
        batch_images(str(source_dir / "*.png"), tmp_path / "limited", width=40)
    assert too_many.value.code == "LIMIT_EXCEEDED"

    monkeypatch.setattr(image_ops, "MAX_IMAGE_BATCH_FILES", 10)
    original_resize = image_ops.resize_image
    calls = 0

    def fail_second(*args: object, **kwargs: object) -> dict[str, object]:
        nonlocal calls
        calls += 1
        if calls == 2:
            raise AerError("CORRUPT_FILE", "simulated failure", "image.batch")
        return original_resize(*args, **kwargs)  # type: ignore[arg-type]

    monkeypatch.setattr(image_ops, "resize_image", fail_second)
    output_dir = tmp_path / "atomic"
    with pytest.raises(AerError):
        batch_images(str(source_dir / "*.png"), output_dir, width=40)
    assert not output_dir.exists()


def test_image_batch_stops_consuming_glob_after_limit(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    consumed = 0

    def unbounded_matches(pattern: str):
        nonlocal consumed
        assert pattern == "*.png"
        for index in range(100):
            consumed += 1
            yield str(tmp_path / f"{index}.png")

    monkeypatch.setattr(image_ops, "MAX_IMAGE_BATCH_FILES", 2)
    monkeypatch.setattr(image_ops.glob, "iglob", unbounded_matches)

    with pytest.raises(AerError) as captured:
        batch_images("*.png", tmp_path / "output", width=40)

    assert captured.value.code == "LIMIT_EXCEEDED"
    assert consumed == 3


def test_image_input_byte_limit_is_checked_before_pillow(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    source = _image(tmp_path / "source.png")
    monkeypatch.setattr(image_ops, "MAX_IMAGE_INPUT_BYTES", source.stat().st_size - 1)
    with pytest.raises(AerError) as limited:
        inspect_image(source)
    assert limited.value.code == "LIMIT_EXCEEDED"


def test_pdf_page_parser_and_merge_extract_split_reopen(tmp_path: Path) -> None:
    first = _pdf(tmp_path / "first.pdf", 2, title="First")
    second = _pdf(tmp_path / "second.pdf", 1, title="Second")
    assert parse_pages("1-2,2,3", 3) == [0, 1, 2]
    with pytest.raises(AerError) as invalid:
        parse_pages("0,4", 3)
    assert invalid.value.code == "INVALID_SELECTOR"

    merged = tmp_path / "merged.pdf"
    merge_result = merge_pdfs([first, second], merged)
    assert merge_result["page_count"] == 3
    assert len(PdfReader(merged).pages) == 3
    assert PdfReader(merged).metadata.title == "First"
    inspected = inspect_pdf(merged, page=2)
    assert inspected["page_count"] == 3
    assert inspected["page"]["number"] == 2

    extracted = tmp_path / "extracted.pdf"
    extract_result = extract_pdf(merged, extracted, pages="1,3")
    assert extract_result["pages"] == [1, 3]
    assert len(PdfReader(extracted).pages) == 2

    split_dir = tmp_path / "split"
    split_result = split_pdf(merged, split_dir)
    assert split_result["count"] == 3
    assert all(len(PdfReader(item["output"]).pages) == 1 for item in split_result["files"])


def test_pdf_encryption_and_input_output_conflicts_are_rejected(tmp_path: Path) -> None:
    encrypted = tmp_path / "encrypted.pdf"
    writer = PdfWriter()
    writer.add_blank_page(width=100, height=100)
    writer.encrypt("password")
    with encrypted.open("wb") as handle:
        writer.write(handle)
    with pytest.raises(AerError) as encryption:
        inspect_pdf(encrypted)
    assert encryption.value.code == "UNSUPPORTED_FORMAT"

    source = _pdf(tmp_path / "same.pdf", 1)
    with pytest.raises(AerError) as conflict:
        extract_pdf(source, source, pages="1")
    assert conflict.value.code == "CONFLICT"


def test_archive_is_deterministic_manifested_and_verifiable(tmp_path: Path) -> None:
    source = tmp_path / "delivery"
    (source / "nested").mkdir(parents=True)
    (source / "a.txt").write_text("alpha", encoding="utf-8")
    (source / "nested" / "b.txt").write_text("beta", encoding="utf-8")
    first = tmp_path / "first.zip"
    second = tmp_path / "second.zip"

    create_archive(source, first)
    create_archive(source, second)
    assert first.read_bytes() == second.read_bytes()
    assert sha256_file(first) == sha256_file(second)
    verified = verify_archive(first)
    assert verified["valid"] is True
    listed = list_archive(first)
    assert listed["entry_count"] == 3
    assert [entry["path"] for entry in listed["entries"]] == [
        "a.txt",
        "nested/b.txt",
        "manifest.json",
    ]
    with zipfile.ZipFile(first) as archive:
        manifest = json.loads(archive.read("manifest.json"))
        assert manifest["deterministic_timestamp"] == "1980-01-01T00:00:00Z"
        assert [item["path"] for item in manifest["files"]] == ["a.txt", "nested/b.txt"]


def test_archive_rejects_reserved_manifest_as_single_input(tmp_path: Path) -> None:
    source = tmp_path / "manifest.json"
    source.write_text("{}\n", encoding="utf-8")
    output = tmp_path / "invalid.zip"

    with pytest.raises(AerError) as conflict:
        create_archive(source, output)

    assert conflict.value.code == "CONFLICT"
    assert not output.exists()


def test_archive_rejects_same_single_file_input_and_output_without_mutation(
    tmp_path: Path,
) -> None:
    source = tmp_path / "source.zip"
    source.write_bytes(b"original archive input")
    before = source.read_bytes()

    with pytest.raises(AerError) as conflict:
        create_archive(source, source)

    assert conflict.value.code == "CONFLICT"
    assert source.read_bytes() == before


def test_archive_rejects_traversal_symlinks_and_hash_tampering(tmp_path: Path) -> None:
    traversal = tmp_path / "traversal.zip"
    with zipfile.ZipFile(traversal, "w") as archive:
        archive.writestr("../escape.txt", "owned")
        archive.writestr("manifest.json", '{"version":1,"files":[]}')
    for operation in (list_archive, verify_archive):
        with pytest.raises(AerError) as unsafe:
            operation(traversal)
        assert unsafe.value.code == "PATH_OUTSIDE_ROOT"

    source = tmp_path / "source"
    source.mkdir()
    (source / "real.txt").write_text("safe")
    if hasattr(os, "symlink"):
        (source / "link.txt").symlink_to(source / "real.txt")
        with pytest.raises(AerError) as symlink:
            create_archive(source, tmp_path / "symlink.zip")
        assert symlink.value.code == "INVALID_ARGUMENT"
        (source / "link.txt").unlink()

        linked_source = tmp_path / "linked-source"
        linked_source.symlink_to(source, target_is_directory=True)
        with pytest.raises(AerError) as top_level_symlink:
            create_archive(linked_source, tmp_path / "linked-source.zip")
        assert top_level_symlink.value.code == "INVALID_ARGUMENT"

        output_target = tmp_path / "output-target.zip"
        output_link = tmp_path / "output-link.zip"
        output_link.symlink_to(output_target)
        with pytest.raises(AerError) as output_symlink:
            create_archive(source, output_link)
        assert output_symlink.value.code == "INVALID_ARGUMENT"
        assert not output_target.exists()

    valid = tmp_path / "valid.zip"
    create_archive(source, valid)
    with zipfile.ZipFile(valid) as archive:
        records = {item.filename: archive.read(item.filename) for item in archive.infolist()}
    tampered = tmp_path / "tampered.zip"
    with zipfile.ZipFile(tampered, "w") as archive:
        for name, data in records.items():
            archive.writestr(name, b"changed" if name == "real.txt" else data)
    with pytest.raises(AerError) as mismatch:
        verify_archive(tampered)
    assert mismatch.value.code == "HASH_MISMATCH"

    duplicate = tmp_path / "duplicate.zip"
    with (
        pytest.warns(UserWarning, match="Duplicate name"),
        zipfile.ZipFile(duplicate, "w") as archive,
    ):
        archive.writestr("same.txt", "first")
        archive.writestr("same.txt", "second")
        archive.writestr("manifest.json", '{"version":1,"files":[]}')
    with pytest.raises(AerError) as duplicate_error:
        verify_archive(duplicate)
    assert duplicate_error.value.code == "CORRUPT_FILE"


@pytest.mark.parametrize(
    "manifest",
    [
        [],
        {"version": 2, "deterministic_timestamp": "1980-01-01T00:00:00Z", "files": []},
        {"version": 1, "deterministic_timestamp": "current-time", "files": []},
    ],
)
def test_archive_verify_rejects_invalid_manifest_contract(tmp_path: Path, manifest: object) -> None:
    target = tmp_path / "invalid-manifest.zip"
    with zipfile.ZipFile(target, "w") as archive:
        archive.writestr("manifest.json", json.dumps(manifest))

    with pytest.raises(AerError) as captured:
        verify_archive(target)

    assert captured.value.code == "CORRUPT_FILE"


def test_conversion_reopens_local_formats_and_reports_missing_dependencies(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    csv_source = tmp_path / "data.csv"
    csv_source.write_text("id,value\n1,10\n", encoding="utf-8")
    workbook_path = tmp_path / "data.xlsx"
    convert_file(csv_source, workbook_path)
    workbook = load_workbook(workbook_path, read_only=True)
    try:
        assert list(workbook.active.values) == [("id", "value"), ("1", "10")]
    finally:
        workbook.close()

    png = _image(tmp_path / "image.png", (20, 10))
    webp = tmp_path / "image.webp"
    convert_file(png, webp)
    with Image.open(webp) as converted:
        assert converted.size == (20, 10)
        assert converted.format == "WEBP"

        monkeypatch.setattr(conversion_ops.shutil, "which", lambda _name: None)
        office = tmp_path / "report.docx"
        Document().save(office)
    with pytest.raises(AerError) as libreoffice:
        convert_file(office, tmp_path / "report.pdf")
    assert libreoffice.value.code == "DEPENDENCY_MISSING"
    assert libreoffice.value.details == {
        "dependency": "libreoffice",
        "capability": "office.to_pdf",
    }

    markdown = tmp_path / "report.md"
    markdown.write_text("# Report")
    with pytest.raises(AerError) as pandoc:
        convert_file(markdown, tmp_path / "report.html")
    assert pandoc.value.code == "DEPENDENCY_MISSING"
    assert pandoc.value.details["dependency"] == "pandoc"


def test_csv_conversion_preserves_formula_like_values_as_text(tmp_path: Path) -> None:
    source = tmp_path / "formula.csv"
    source.write_text("value\n=2+2\n", encoding="utf-8")
    output = tmp_path / "formula.xlsx"

    convert_file(source, output)

    workbook = load_workbook(output, data_only=False)
    try:
        cell = workbook.active["A2"]
        assert cell.value == "=2+2"
        assert cell.data_type == "s"
        assert cell.quotePrefix is True
    finally:
        workbook.close()


def test_csv_to_xlsx_enforces_cumulative_cell_limit_for_ragged_rows(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    source = tmp_path / "ragged.csv"
    source.write_text("a,b,c,d\n1,2\n", encoding="utf-8")
    output = tmp_path / "ragged.xlsx"
    monkeypatch.setattr(conversion_ops, "MAX_TABULAR_CELLS", 5)

    with pytest.raises(AerError) as exceeded:
        convert_file(source, output)

    assert exceeded.value.code == "LIMIT_EXCEEDED"
    assert exceeded.value.details["cells"] == 6
    assert exceeded.value.details["cell_limit"] == 5
    assert not output.exists()


def test_xlsx_to_tsv_uses_tab_delimiter(tmp_path: Path) -> None:
    source = tmp_path / "table.xlsx"
    book = Workbook()
    book.active.append(["id", "value"])
    book.active.append(["A,1", "tabular"])
    book.save(source)
    book.close()
    output = tmp_path / "table.tsv"

    convert_file(source, output)

    assert output.read_bytes() == b"id\tvalue\r\nA,1\ttabular\r\n"
