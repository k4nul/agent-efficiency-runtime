"""Safe, non-evaluating tabular data loader and query pipeline."""

from __future__ import annotations

import csv
import io
import json
import math
import os
import re
import tempfile
import time
from collections import Counter, defaultdict
from collections.abc import Iterable, Mapping, Sequence
from dataclasses import dataclass, field
from datetime import date, datetime
from datetime import time as datetime_time
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Protocol

from aer.errors import AerError
from aer.limits import MAX_STDIN_BYTES, MAX_TABULAR_CELLS, PREVIEW_ROWS
from aer.paths import prepare_output_path
from aer.protocol import Metrics, success
from aer.zip_safety import enforce_zip_expansion_limits

Row = dict[str, object]

_SYMBOLIC_FILTER_RE = re.compile(
    r"^\s*(?P<column>.+?)\s*(?P<operator>>=|<=|==|!=|>|<)\s*(?P<value>.*?)\s*$"
)
_WORD_FILTER_RE = re.compile(
    r"^\s*(?P<column>.+?)\s+"
    r"(?P<operator>starts_with|ends_with|contains|is_null|not_null|in)"
    r"(?:\s+(?P<value>.*?))?\s*$",
    re.IGNORECASE,
)
_INTEGER_RE = re.compile(r"^-?(?:0|[1-9]\d*)$")
_DECIMAL_RE = re.compile(r"^-?(?:(?:0|[1-9]\d*)\.\d+|(?:0|[1-9]\d*)[eE][+-]?\d+)$")
_AGGREGATE_RE = re.compile(
    r"^\s*(?P<operation>count|sum|average|avg|min|max|null_count)"
    r"(?:\s*[:(]\s*(?P<column>[^):]+?)\s*\)?)?"
    r"(?:\s+(?:as\s+)?(?P<alias>[A-Za-z_][\w.-]*))?\s*$",
    re.IGNORECASE,
)
_MAX_ROWS = 1_000_000


class StoredObject(Protocol):
    @property
    def ref(self) -> str: ...


class ContentStore(Protocol):
    def put_bytes(
        self,
        data: bytes,
        *,
        filename: str | None = None,
        mime_type: str | None = None,
        source: Mapping[str, object] | None = None,
    ) -> StoredObject: ...


@dataclass(frozen=True, slots=True)
class FilterExpression:
    column: str
    operator: str
    value: object = None

    def matches(self, row: Mapping[str, object]) -> bool:
        actual = row.get(self.column)
        if self.operator == "is_null":
            return actual is None
        if self.operator == "not_null":
            return actual is not None
        if self.operator == "in":
            assert isinstance(self.value, tuple)
            return any(_values_equal(actual, candidate) for candidate in self.value)
        if self.operator == "==":
            return _values_equal(actual, self.value)
        if self.operator == "!=":
            return not _values_equal(actual, self.value)
        if self.operator in {">", ">=", "<", "<="}:
            if actual is None or self.value is None:
                return False
            comparison = _compare_values(actual, self.value)
            return {
                ">": comparison > 0,
                ">=": comparison >= 0,
                "<": comparison < 0,
                "<=": comparison <= 0,
            }[self.operator]
        if self.operator == "contains":
            if actual is None:
                return False
            if isinstance(actual, (list, tuple, set)):
                return any(_values_equal(item, self.value) for item in actual)
            return str(self.value) in str(actual)
        if self.operator == "starts_with":
            return actual is not None and str(actual).startswith(str(self.value))
        if self.operator == "ends_with":
            return actual is not None and str(actual).endswith(str(self.value))
        raise AssertionError(f"Unhandled filter operator: {self.operator}")


@dataclass(frozen=True, slots=True)
class AggregateSpec:
    operation: str
    column: str | None = None
    alias: str | None = None

    @property
    def output_name(self) -> str:
        if self.alias:
            return self.alias
        return "count" if self.operation == "count" else f"{self.operation}_{self.column}"


@dataclass(slots=True)
class DataQueryResult:
    source_rows: int
    matched_rows: int
    result_rows: int
    columns: list[str]
    preview: list[Row]
    result_ref: str | None
    output: str | None
    duplicate_rows: int = 0
    duration_ms: int = 0
    bytes_read: int = 0
    bytes_written: int = 0
    warnings: list[dict[str, Any] | str] = field(default_factory=list)

    def to_dict(self) -> dict[str, object]:
        return {
            "source_rows": self.source_rows,
            "matched_rows": self.matched_rows,
            "result_rows": self.result_rows,
            "columns": self.columns,
            "preview": self.preview[:PREVIEW_ROWS],
            "result_ref": self.result_ref,
            "output": self.output,
            "duplicate_rows": self.duplicate_rows,
        }


@dataclass(slots=True)
class _LoadedData:
    rows: list[Row]
    columns: list[str]
    bytes_read: int


def _unquote_column(value: str) -> str:
    value = value.strip()
    if len(value) >= 2 and value[0] == value[-1] and value[0] in {'"', "'"}:
        return value[1:-1]
    return value


def _parse_scalar(value: str) -> object:
    value = value.strip()
    if not value:
        return ""
    if len(value) >= 2 and value[0] == value[-1] and value[0] in {'"', "'"}:
        return value[1:-1]
    lowered = value.casefold()
    if lowered in {"null", "none"}:
        return None
    if lowered == "true":
        return True
    if lowered == "false":
        return False
    if _INTEGER_RE.fullmatch(value):
        return int(value)
    if _DECIMAL_RE.fullmatch(value):
        return Decimal(value)
    return value


def parse_filter(expression: str) -> FilterExpression:
    """Parse the documented filter grammar without evaluating code."""

    if len(expression) > 4096:
        raise AerError(
            "LIMIT_EXCEEDED",
            "Filter expression exceeds 4096 characters.",
            operation="data.query",
            target="where",
        )
    match = _SYMBOLIC_FILTER_RE.fullmatch(expression) or _WORD_FILTER_RE.fullmatch(expression)
    if match is None:
        raise AerError(
            "INVALID_ARGUMENT",
            "Invalid filter. Use: COLUMN OP VALUE, COLUMN is_null, or COLUMN not_null.",
            operation="data.query",
            target="where",
            suggested_action=(
                "Use one of: ==, !=, >, >=, <, <=, contains, starts_with, "
                "ends_with, is_null, not_null, in"
            ),
        )
    column = _unquote_column(match.group("column"))
    operator = match.group("operator").casefold()
    raw_value = match.groupdict().get("value")
    if not column:
        raise AerError(
            "INVALID_ARGUMENT", "Filter column is empty.", operation="data.query", target="where"
        )
    if operator in {"is_null", "not_null"}:
        if raw_value not in {None, ""}:
            raise AerError(
                "INVALID_ARGUMENT",
                f"{operator} does not accept a value.",
                operation="data.query",
                target="where",
            )
        return FilterExpression(column, operator)
    if raw_value is None or not raw_value.strip():
        raise AerError(
            "INVALID_ARGUMENT",
            f"{operator} requires a value.",
            operation="data.query",
            target="where",
        )
    if operator == "in":
        raw_value = raw_value.strip()
        if raw_value.startswith("["):
            try:
                parsed = json.loads(raw_value)
            except json.JSONDecodeError as exc:
                raise AerError(
                    "INVALID_ARGUMENT",
                    f"Invalid JSON list for 'in': {exc.msg}",
                    operation="data.query",
                    target="where",
                ) from exc
            if not isinstance(parsed, list):
                raise AerError(
                    "INVALID_ARGUMENT",
                    "The 'in' value must be a JSON list or comma-separated values.",
                    operation="data.query",
                    target="where",
                )
            values = tuple(_normalize_value(item) for item in parsed)
        else:
            try:
                parts = next(csv.reader([raw_value], skipinitialspace=True))
            except csv.Error as exc:
                raise AerError(
                    "INVALID_ARGUMENT",
                    f"Invalid comma-separated 'in' values: {exc}",
                    operation="data.query",
                    target="where",
                ) from exc
            values = tuple(_parse_scalar(part) for part in parts)
        return FilterExpression(column, operator, values)
    return FilterExpression(column, operator, _parse_scalar(raw_value))


def parse_aggregate(expression: str) -> AggregateSpec:
    """Parse ``count`` or ``sum:column [alias]`` style aggregate specs."""

    match = _AGGREGATE_RE.fullmatch(expression)
    if match is None:
        raise AerError(
            "INVALID_ARGUMENT",
            "Invalid aggregate. Use count or OP:COLUMN [ALIAS].",
            operation="data.query",
            target="aggregate",
            suggested_action="Supported aggregates: count, sum, average, min, max, null_count",
        )
    operation = match.group("operation").casefold()
    if operation == "avg":
        operation = "average"
    column = match.group("column")
    alias = match.group("alias")
    if operation != "count" and not column:
        raise AerError(
            "INVALID_ARGUMENT",
            f"Aggregate {operation} requires a column.",
            operation="data.query",
            target="aggregate",
        )
    if operation == "count" and column:
        raise AerError(
            "INVALID_ARGUMENT",
            "count does not accept a column; use count alone.",
            operation="data.query",
            target="aggregate",
        )
    return AggregateSpec(operation, column.strip() if column else None, alias)


def _normalize_value(value: object) -> object:
    if isinstance(value, (datetime, date, datetime_time)):
        return value.isoformat()
    if isinstance(value, Decimal):
        if value == value.to_integral_value():
            return int(value)
        return float(value)
    if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
        return str(value)
    if isinstance(value, Mapping):
        return {str(key): _normalize_value(item) for key, item in value.items()}
    if isinstance(value, (list, tuple)):
        return [_normalize_value(item) for item in value]
    return value


def _infer_csv_value(value: str) -> object:
    if value == "":
        return None
    return _normalize_value(_parse_scalar(value))


def _ensure_input(path: Path) -> int:
    if not path.is_file():
        raise AerError(
            "NOT_FOUND", "Data source does not exist.", operation="data.query", target=str(path)
        )
    size = path.stat().st_size
    if size > MAX_STDIN_BYTES:
        raise AerError(
            "LIMIT_EXCEEDED",
            f"Data source exceeds the {MAX_STDIN_BYTES}-byte input limit.",
            operation="data.query",
            target=str(path),
        )
    return size


def _validate_columns(columns: Sequence[object], *, target: str) -> list[str]:
    names = [str(item).strip() if item is not None else "" for item in columns]
    if not names or any(not name for name in names):
        raise AerError(
            "INVALID_SPEC",
            "Tabular data must have non-empty column names.",
            operation="data.query",
            target=target,
        )
    duplicates = sorted(name for name, count in Counter(names).items() if count > 1)
    if duplicates:
        raise AerError(
            "INVALID_SPEC",
            "Tabular data has duplicate column names.",
            operation="data.query",
            target=target,
            details={"duplicates": duplicates},
        )
    return names


def _enforce_table_shape(row_count: int, column_count: int, *, target: str) -> None:
    cells = row_count * column_count
    if cells > MAX_TABULAR_CELLS:
        raise AerError(
            "LIMIT_EXCEEDED",
            f"Data source exceeds the {MAX_TABULAR_CELLS}-cell limit.",
            operation="data.query",
            target=target,
            details={
                "rows": row_count,
                "columns": column_count,
                "cells": cells,
                "limit": MAX_TABULAR_CELLS,
            },
        )


def _load_delimited(path: Path, *, delimiter: str, size: int) -> _LoadedData:
    try:
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.reader(handle, delimiter=delimiter)
            try:
                raw_header = next(reader)
            except StopIteration:
                raise AerError(
                    "INVALID_SPEC",
                    "Delimited data is empty.",
                    operation="data.query",
                    target=str(path),
                ) from None
            columns = _validate_columns(raw_header, target=str(path))
            rows: list[Row] = []
            for line_number, values in enumerate(reader, start=2):
                if not values or all(value == "" for value in values):
                    continue
                if len(values) != len(columns):
                    raise AerError(
                        "INVALID_SPEC",
                        f"Row {line_number} has {len(values)} values; expected {len(columns)}.",
                        operation="data.query",
                        target=f"{path}:{line_number}",
                    )
                rows.append(dict(zip(columns, map(_infer_csv_value, values), strict=True)))
                _enforce_table_shape(len(rows), len(columns), target=str(path))
                if len(rows) > _MAX_ROWS:
                    raise AerError(
                        "LIMIT_EXCEEDED",
                        f"Data source exceeds the {_MAX_ROWS}-row limit.",
                        operation="data.query",
                        target=str(path),
                    )
    except UnicodeDecodeError as exc:
        raise AerError(
            "CORRUPT_FILE",
            "Delimited data is not valid UTF-8.",
            operation="data.query",
            target=str(path),
        ) from exc
    except csv.Error as exc:
        raise AerError(
            "CORRUPT_FILE",
            f"Delimited data could not be parsed: {exc}",
            operation="data.query",
            target=str(path),
        ) from exc
    return _LoadedData(rows, columns, size)


def _json_rows(value: object, *, target: str) -> tuple[list[Row], list[str]]:
    if not isinstance(value, list) or any(not isinstance(item, dict) for item in value):
        raise AerError(
            "INVALID_SPEC",
            "JSON data must be an array of objects.",
            operation="data.query",
            target=target,
        )
    rows: list[Row] = []
    columns: list[str] = []
    seen: set[str] = set()
    for source_row in value:
        assert isinstance(source_row, dict)
        row = {str(key): _normalize_value(item) for key, item in source_row.items()}
        rows.append(row)
        for column in row:
            if column not in seen:
                columns.append(column)
                seen.add(column)
        _enforce_table_shape(len(rows), len(columns), target=target)
    return rows, columns


def _load_json(path: Path, *, size: int) -> _LoadedData:
    try:
        value = json.loads(path.read_text(encoding="utf-8-sig"))
    except (UnicodeDecodeError, json.JSONDecodeError) as exc:
        message = exc.msg if isinstance(exc, json.JSONDecodeError) else "invalid UTF-8"
        raise AerError(
            "CORRUPT_FILE",
            f"JSON data could not be parsed: {message}",
            operation="data.query",
            target=str(path),
        ) from exc
    rows, columns = _json_rows(value, target=str(path))
    if len(rows) > _MAX_ROWS:
        raise AerError(
            "LIMIT_EXCEEDED",
            f"Data source exceeds the {_MAX_ROWS}-row limit.",
            operation="data.query",
            target=str(path),
        )
    return _LoadedData(rows, columns, size)


def _load_jsonl(path: Path, *, size: int) -> _LoadedData:
    rows: list[Row] = []
    columns: list[str] = []
    seen: set[str] = set()
    try:
        with path.open("r", encoding="utf-8-sig") as handle:
            for line_number, line in enumerate(handle, start=1):
                if not line.strip():
                    continue
                try:
                    value = json.loads(line)
                except json.JSONDecodeError as exc:
                    raise AerError(
                        "CORRUPT_FILE",
                        f"JSONL line {line_number} could not be parsed: {exc.msg}",
                        operation="data.query",
                        target=f"{path}:{line_number}",
                    ) from exc
                if not isinstance(value, dict):
                    raise AerError(
                        "INVALID_SPEC",
                        f"JSONL line {line_number} is not an object.",
                        operation="data.query",
                        target=f"{path}:{line_number}",
                    )
                row = {str(key): _normalize_value(item) for key, item in value.items()}
                rows.append(row)
                for column in row:
                    if column not in seen:
                        columns.append(column)
                        seen.add(column)
                _enforce_table_shape(len(rows), len(columns), target=str(path))
                if len(rows) > _MAX_ROWS:
                    raise AerError(
                        "LIMIT_EXCEEDED",
                        f"Data source exceeds the {_MAX_ROWS}-row limit.",
                        operation="data.query",
                        target=str(path),
                    )
    except UnicodeDecodeError as exc:
        raise AerError(
            "CORRUPT_FILE",
            "JSONL data is not valid UTF-8.",
            operation="data.query",
            target=str(path),
        ) from exc
    return _LoadedData(rows, columns, size)


def _load_xlsx(path: Path, *, size: int, sheet: str | None) -> _LoadedData:
    enforce_zip_expansion_limits(path, operation="data.query", target=str(path))
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=False)
    except Exception as exc:
        raise AerError(
            "CORRUPT_FILE",
            f"Workbook could not be opened: {exc}",
            operation="data.query",
            target=str(path),
        ) from exc
    try:
        if sheet is not None and sheet not in workbook.sheetnames:
            raise AerError(
                "NOT_FOUND",
                f"Workbook sheet was not found: {sheet}",
                operation="data.query",
                target=sheet,
                details={"available": workbook.sheetnames},
            )
        worksheet = workbook[sheet] if sheet is not None else workbook.active
        iterator = worksheet.iter_rows(values_only=True)
        try:
            raw_header = next(iterator)
        except StopIteration:
            raise AerError(
                "INVALID_SPEC",
                "Workbook sheet is empty.",
                operation="data.query",
                target=worksheet.title,
            ) from None
        columns = _validate_columns(raw_header, target=f"{path}:{worksheet.title}")
        rows: list[Row] = []
        for row_number, values in enumerate(iterator, start=2):
            normalized = [_normalize_value(value) for value in values]
            if len(normalized) < len(columns):
                normalized.extend([None] * (len(columns) - len(normalized)))
            if len(normalized) > len(columns) and any(
                value is not None for value in normalized[len(columns) :]
            ):
                raise AerError(
                    "INVALID_SPEC",
                    f"Workbook row {row_number} has values beyond the header.",
                    operation="data.query",
                    target=f"{worksheet.title}:{row_number}",
                )
            rows.append(dict(zip(columns, normalized[: len(columns)], strict=True)))
            _enforce_table_shape(len(rows), len(columns), target=str(path))
            if len(rows) > _MAX_ROWS:
                raise AerError(
                    "LIMIT_EXCEEDED",
                    f"Data source exceeds the {_MAX_ROWS}-row limit.",
                    operation="data.query",
                    target=str(path),
                )
    finally:
        workbook.close()
    return _LoadedData(rows, columns, size)


def _load_data(path: Path, *, sheet: str | None) -> _LoadedData:
    size = _ensure_input(path)
    suffix = path.suffix.casefold()
    if suffix == ".csv":
        return _load_delimited(path, delimiter=",", size=size)
    if suffix == ".tsv":
        return _load_delimited(path, delimiter="\t", size=size)
    if suffix == ".json":
        return _load_json(path, size=size)
    if suffix in {".jsonl", ".ndjson"}:
        return _load_jsonl(path, size=size)
    if suffix == ".xlsx":
        return _load_xlsx(path, size=size, sheet=sheet)
    raise AerError(
        "UNSUPPORTED_FORMAT",
        f"Unsupported data format: {suffix or '(none)'}",
        operation="data.query",
        target=str(path),
        suggested_action="Use CSV, TSV, JSON array, JSONL, or XLSX.",
    )


def _as_decimal(value: object) -> Decimal | None:
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, (int, float, Decimal)):
        try:
            return Decimal(str(value))
        except InvalidOperation:
            return None
    if isinstance(value, str):
        try:
            return Decimal(value.strip())
        except InvalidOperation:
            return None
    return None


def _values_equal(left: object, right: object) -> bool:
    if left is None or right is None:
        return left is right
    left_number = _as_decimal(left)
    right_number = _as_decimal(right)
    if left_number is not None and right_number is not None:
        return left_number == right_number
    if isinstance(left, bool) or isinstance(right, bool):
        return left is right
    return str(left) == str(right)


def _compare_values(left: object, right: object) -> int:
    if left is None or right is None:
        return (left is not None) - (right is not None)
    left_number = _as_decimal(left)
    right_number = _as_decimal(right)
    if left_number is not None and right_number is not None:
        return (left_number > right_number) - (left_number < right_number)
    left_text, right_text = str(left), str(right)
    return (left_text > right_text) - (left_text < right_text)


def _require_columns(requested: Iterable[str], available: Sequence[str], *, target: str) -> None:
    missing = sorted(set(requested) - set(available))
    if missing:
        raise AerError(
            "INVALID_ARGUMENT",
            "Query references columns that do not exist.",
            operation="data.query",
            target=target,
            details={"missing": missing, "available": list(available)},
        )


def _freeze(value: object) -> object:
    if isinstance(value, dict):
        return tuple(sorted((str(key), _freeze(item)) for key, item in value.items()))
    if isinstance(value, (list, tuple)):
        return tuple(_freeze(item) for item in value)
    if isinstance(value, set):
        return tuple(sorted((_freeze(item) for item in value), key=repr))
    return value


def _row_key(row: Mapping[str, object], columns: Sequence[str]) -> tuple[object, ...]:
    return tuple(_freeze(row.get(column)) for column in columns)


def _deduplicate(rows: Sequence[Row], columns: Sequence[str]) -> list[Row]:
    seen: set[tuple[object, ...]] = set()
    result: list[Row] = []
    for row in rows:
        key = _row_key(row, columns)
        if key not in seen:
            seen.add(key)
            result.append(row)
    return result


def _duplicate_rows(rows: Sequence[Row], columns: Sequence[str]) -> list[Row]:
    counts = Counter(_row_key(row, columns) for row in rows)
    return [row for row in rows if counts[_row_key(row, columns)] > 1]


def _aggregate_value(rows: Sequence[Row], spec: AggregateSpec) -> object:
    if spec.operation == "count":
        return len(rows)
    assert spec.column is not None
    values = [row.get(spec.column) for row in rows]
    if spec.operation == "null_count":
        return sum(value is None for value in values)
    non_null = [value for value in values if value is not None]
    if spec.operation in {"sum", "average"}:
        numeric: list[Decimal] = []
        for value in non_null:
            number = _as_decimal(value)
            if number is None:
                raise AerError(
                    "INVALID_ARGUMENT",
                    f"Aggregate {spec.operation} requires numeric values in {spec.column}.",
                    operation="data.query",
                    target=spec.column,
                )
            numeric.append(number)
        if spec.operation == "sum":
            return _normalize_value(sum(numeric, Decimal(0)))
        return _normalize_value(sum(numeric, Decimal(0)) / len(numeric)) if numeric else None
    if not non_null:
        return None
    if spec.operation == "min":
        return min(non_null, key=_sort_value)
    if spec.operation == "max":
        return max(non_null, key=_sort_value)
    raise AssertionError(f"Unhandled aggregate operation: {spec.operation}")


def _aggregate_rows(
    rows: Sequence[Row], group_by: Sequence[str], aggregates: Sequence[AggregateSpec]
) -> list[Row]:
    grouped: dict[tuple[object, ...], list[Row]] = defaultdict(list)
    if group_by:
        for row in rows:
            grouped[_row_key(row, group_by)].append(row)
    else:
        grouped[()] = list(rows)
    result: list[Row] = []
    for key, members in grouped.items():
        output: Row = {
            column: members[0].get(column) if members else key[index]
            for index, column in enumerate(group_by)
        }
        for aggregate in aggregates:
            output[aggregate.output_name] = _aggregate_value(members, aggregate)
        result.append(output)
    return result


def _sort_value(value: object) -> tuple[int, object]:
    if value is None:
        return (2, "")
    number = _as_decimal(value)
    if number is not None:
        return (0, number)
    return (1, str(value).casefold())


def _parse_name_list(value: str | Sequence[str] | None) -> list[str] | None:
    if value is None:
        return None
    raw = [value] if isinstance(value, str) else list(value)
    names = [part.strip() for item in raw for part in item.split(",") if part.strip()]
    return names


def _parse_rename(
    rename: Mapping[str, str] | Sequence[str] | None,
) -> dict[str, str]:
    if rename is None:
        return {}
    result: dict[str, str]
    if isinstance(rename, Mapping):
        result = {str(source): str(target) for source, target in rename.items()}
    else:
        result = {}
        for item in rename:
            if ":" not in item:
                raise AerError(
                    "INVALID_ARGUMENT",
                    "Rename entries must use OLD:NEW.",
                    operation="data.query",
                    target="rename",
                )
            source, target = item.split(":", 1)
            result[source.strip()] = target.strip()
    if any(not source or not target for source, target in result.items()):
        raise AerError(
            "INVALID_ARGUMENT",
            "Rename source and target names must be non-empty.",
            operation="data.query",
            target="rename",
        )
    return result


def _csv_value(value: object) -> object:
    if value is None:
        return ""
    if isinstance(value, (dict, list, tuple)):
        return json.dumps(value, ensure_ascii=False, separators=(",", ":"), sort_keys=True)
    return value


def _serialize_text(rows: Sequence[Row], columns: Sequence[str], suffix: str) -> bytes:
    if suffix == ".json":
        return json.dumps(rows, ensure_ascii=False, separators=(",", ":")).encode("utf-8") + b"\n"
    if suffix in {".jsonl", ".ndjson"}:
        return (
            "".join(
                json.dumps(row, ensure_ascii=False, separators=(",", ":")) + "\n" for row in rows
            )
        ).encode("utf-8")
    delimiter = "\t" if suffix == ".tsv" else ","
    buffer = io.StringIO(newline="")
    writer = csv.DictWriter(
        buffer,
        fieldnames=list(columns),
        delimiter=delimiter,
        lineterminator="\n",
    )
    writer.writeheader()
    for row in rows:
        writer.writerow({column: _csv_value(row.get(column)) for column in columns})
    return buffer.getvalue().encode("utf-8")


def _atomic_write_bytes(output: Path, data: bytes) -> None:
    output.parent.mkdir(parents=True, exist_ok=True)
    descriptor, temporary_name = tempfile.mkstemp(prefix=f".{output.name}.", dir=output.parent)
    temporary = Path(temporary_name)
    try:
        os.fchmod(descriptor, 0o600)
        with os.fdopen(descriptor, "wb") as handle:
            handle.write(data)
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(temporary, output)
    except BaseException:
        temporary.unlink(missing_ok=True)
        raise


def _write_xlsx(output: Path, rows: Sequence[Row], columns: Sequence[str]) -> None:
    from openpyxl import Workbook
    from openpyxl.cell import WriteOnlyCell

    def safe_cell(worksheet: Any, value: object) -> Any:
        cell = WriteOnlyCell(worksheet, value=value)
        if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
            cell.data_type = "s"
            cell.quotePrefix = True
        return cell

    output.parent.mkdir(parents=True, exist_ok=True)
    descriptor, temporary_name = tempfile.mkstemp(
        prefix=f".{output.name}.", suffix=".xlsx", dir=output.parent
    )
    os.close(descriptor)
    temporary = Path(temporary_name)
    try:
        workbook = Workbook(write_only=True)
        worksheet = workbook.create_sheet("Query")
        worksheet.append([safe_cell(worksheet, value) for value in columns])
        for row in rows:
            worksheet.append(
                [safe_cell(worksheet, _csv_value(row.get(column))) for column in columns]
            )
        workbook.save(temporary)
        os.chmod(temporary, 0o600)
        os.replace(temporary, output)
    except BaseException:
        temporary.unlink(missing_ok=True)
        raise


def _write_result(output: Path, rows: Sequence[Row], columns: Sequence[str]) -> int:
    suffix = output.suffix.casefold()
    if suffix == ".xlsx":
        _write_xlsx(output, rows, columns)
    elif suffix in {".csv", ".tsv", ".json", ".jsonl", ".ndjson"}:
        _atomic_write_bytes(output, _serialize_text(rows, columns, suffix))
    else:
        raise AerError(
            "UNSUPPORTED_FORMAT",
            f"Unsupported query output format: {suffix or '(none)'}",
            operation="data.query",
            target=str(output),
            suggested_action="Use CSV, TSV, JSON, JSONL, or XLSX.",
        )
    return output.stat().st_size


def query_data(
    source: str | Path,
    *,
    sheet: str | None = None,
    where: str | Sequence[str] = (),
    select: str | Sequence[str] | None = None,
    rename: Mapping[str, str] | Sequence[str] | None = None,
    sort: str | None = None,
    descending: bool = False,
    limit: int | None = None,
    offset: int = 0,
    unique: bool | str | Sequence[str] = False,
    group_by: str | Sequence[str] | None = None,
    aggregates: Sequence[str | AggregateSpec] = (),
    duplicate_columns: str | Sequence[str] | None = None,
    output: str | Path | None = None,
    store: ContentStore | None = None,
) -> DataQueryResult:
    """Load, query, preview, and optionally persist a tabular result.

    The operation order is filter, duplicate/unique handling, aggregate or select,
    rename, sort, offset, and limit. ``matched_rows`` is measured before pagination.
    """

    started = time.monotonic()
    requested_source = Path(source).expanduser()
    if requested_source.is_symlink():
        raise AerError(
            "INVALID_ARGUMENT",
            "Data source cannot be a symbolic link.",
            operation="data.query",
            target=str(requested_source),
        )
    source_path = requested_source.resolve()
    output_path = (
        prepare_output_path(Path(output), operation="data.query") if output is not None else None
    )
    if output_path is not None and output_path == source_path:
        raise AerError(
            "CONFLICT",
            "Query output must differ from the input path.",
            operation="data.query",
            target=str(output_path),
        )
    if offset < 0 or (limit is not None and limit < 0):
        raise AerError(
            "INVALID_ARGUMENT",
            "Offset and limit must be non-negative.",
            operation="data.query",
            target="pagination",
        )

    loaded = _load_data(source_path, sheet=sheet)
    where_expressions = [where] if isinstance(where, str) else where
    filters = [parse_filter(expression) for expression in where_expressions]
    _require_columns((item.column for item in filters), loaded.columns, target="where")
    filtered = [row for row in loaded.rows if all(item.matches(row) for item in filters)]

    duplicate_names = _parse_name_list(duplicate_columns)
    duplicate_count = 0
    if duplicate_names is not None:
        if not duplicate_names:
            duplicate_names = list(loaded.columns)
        _require_columns(duplicate_names, loaded.columns, target="duplicates")
        filtered = _duplicate_rows(filtered, duplicate_names)
        duplicate_count = len(filtered)

    if unique:
        unique_names: list[str]
        if unique is True:
            unique_names = list(loaded.columns)
        else:
            parsed_unique_names = _parse_name_list(unique)
            assert parsed_unique_names is not None
            unique_names = parsed_unique_names
        _require_columns(unique_names, loaded.columns, target="unique")
        filtered = _deduplicate(filtered, unique_names)

    matched_rows = len(filtered)
    parsed_aggregates = [
        item if isinstance(item, AggregateSpec) else parse_aggregate(item) for item in aggregates
    ]
    group_names = _parse_name_list(group_by) or []
    selected = _parse_name_list(select)
    if parsed_aggregates:
        aggregate_columns = [item.column for item in parsed_aggregates if item.column is not None]
        _require_columns([*group_names, *aggregate_columns], loaded.columns, target="aggregate")
        rows = _aggregate_rows(filtered, group_names, parsed_aggregates)
        columns = [*group_names, *(item.output_name for item in parsed_aggregates)]
        if len(set(columns)) != len(columns):
            raise AerError(
                "CONFLICT",
                "Aggregate aliases create duplicate output columns.",
                operation="data.query",
                target="aggregate",
            )
    else:
        if group_names:
            raise AerError(
                "INVALID_ARGUMENT",
                "group_by requires at least one aggregate.",
                operation="data.query",
                target="group_by",
            )
        columns = selected if selected is not None else list(loaded.columns)
        _require_columns(columns, loaded.columns, target="select")
        rows = [{column: row.get(column) for column in columns} for row in filtered]

    rename_map = _parse_rename(rename)
    _require_columns(rename_map, columns, target="rename")
    renamed_columns = [rename_map.get(column, column) for column in columns]
    if len(set(renamed_columns)) != len(renamed_columns):
        raise AerError(
            "CONFLICT",
            "Rename would create duplicate output columns.",
            operation="data.query",
            target="rename",
        )
    if rename_map:
        rows = [
            {rename_map.get(column, column): row.get(column) for column in columns} for row in rows
        ]
    columns = renamed_columns

    if sort is not None:
        if sort not in columns:
            raise AerError(
                "INVALID_ARGUMENT",
                f"Sort column does not exist: {sort}",
                operation="data.query",
                target="sort",
                details={"available": columns},
            )
        rows.sort(key=lambda row: _sort_value(row.get(sort)), reverse=descending)
    paged_rows = rows[offset:]
    if limit is not None:
        paged_rows = paged_rows[:limit]

    bytes_written = 0
    if output_path is not None:
        bytes_written = _write_result(output_path, paged_rows, columns)

    result_ref: str | None = None
    if store is not None:
        if output_path is not None:
            stored_data = output_path.read_bytes()
            stored_name = output_path.name
            mime_type = {
                ".csv": "text/csv",
                ".tsv": "text/tab-separated-values",
                ".json": "application/json",
                ".jsonl": "application/x-ndjson",
                ".ndjson": "application/x-ndjson",
                ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            }[output_path.suffix.casefold()]
        else:
            stored_data = _serialize_text(paged_rows, columns, ".jsonl")
            stored_name = f"{source_path.stem}.query.jsonl"
            mime_type = "application/x-ndjson"
        result_ref = store.put_bytes(
            stored_data,
            filename=stored_name,
            mime_type=mime_type,
            source={"operation": "data.query", "input": str(source_path)},
        ).ref

    duration_ms = round((time.monotonic() - started) * 1000)
    return DataQueryResult(
        source_rows=len(loaded.rows),
        matched_rows=matched_rows,
        result_rows=len(paged_rows),
        columns=columns,
        preview=paged_rows[:PREVIEW_ROWS],
        result_ref=result_ref,
        output=str(output_path) if output_path is not None else None,
        duplicate_rows=duplicate_count,
        duration_ms=duration_ms,
        bytes_read=loaded.bytes_read,
        bytes_written=bytes_written,
    )


def data_response(result: DataQueryResult) -> dict[str, object]:
    """Convert a query result to the repository-wide response protocol."""

    return success(
        "data.query",
        result.to_dict(),
        artifacts=(
            [{"ref": result.result_ref, "role": "query_result"}] if result.result_ref else []
        ),
        warnings=result.warnings,
        metrics=Metrics(
            duration_ms=result.duration_ms,
            bytes_read=result.bytes_read,
            bytes_written=result.bytes_written,
        ),
    )
