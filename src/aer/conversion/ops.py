from __future__ import annotations

import csv
import io
import re
import shutil
import subprocess
import tempfile
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from PIL import Image, ImageOps

from aer.errors import AerError
from aer.hashing import sha256_file
from aer.limits import (
    MAX_IMAGE_PIXELS,
    MAX_TABULAR_CELLS,
    MAX_TABULAR_FILE_BYTES,
    MAX_TABULAR_ROWS,
    MAX_TEXT_FILE_BYTES,
    MAX_ZIP_UNCOMPRESSED_BYTES,
)
from aer.paths import atomic_binary_writer, ensure_regular_input, prepare_output_path
from aer.zip_safety import enforce_zip_expansion_limits, inspect_zip_active_content


def _dependency(name: str, capability: str, target: Path) -> AerError:
    return AerError(
        "DEPENDENCY_MISSING",
        f"{name} is required for this conversion.",
        "artifact.convert",
        str(target),
        {"dependency": name, "capability": capability},
    )


def _copy_bounded(source: Path, output: Path) -> None:
    size = source.stat().st_size
    if size > MAX_ZIP_UNCOMPRESSED_BYTES:
        raise AerError(
            "LIMIT_EXCEEDED",
            "Converted output exceeds the file-size safety limit.",
            "artifact.convert",
            str(source),
            {"bytes": size, "limit": MAX_ZIP_UNCOMPRESSED_BYTES},
        )
    with source.open("rb") as input_handle, atomic_binary_writer(output) as output_handle:
        shutil.copyfileobj(input_handle, output_handle, length=1024 * 1024)


def _office_to_pdf(source: Path, output: Path) -> None:
    enforce_zip_expansion_limits(source, operation="artifact.convert", target=str(source))
    active = inspect_zip_active_content(source, operation="artifact.convert", target=str(source))
    if active.external_links or active.active_parts:
        raise AerError(
            "UNSUPPORTED_FORMAT",
            "Office conversion rejects external relationships, macros, and executable parts.",
            "artifact.convert",
            str(source),
            {
                "external_links": list(active.external_links),
                "active_parts": list(active.active_parts),
            },
            "Remove active content before conversion.",
        )
    executable = shutil.which("libreoffice") or shutil.which("soffice")
    if not executable:
        raise _dependency("libreoffice", "office.to_pdf", source)
    with tempfile.TemporaryDirectory(prefix="aer-convert-") as directory:
        profile = Path(directory) / "lo-profile"
        profile.mkdir(mode=0o700)
        try:
            completed = subprocess.run(
                [
                    executable,
                    "--headless",
                    "--nologo",
                    "--nodefault",
                    "--nolockcheck",
                    "--norestore",
                    "--safe-mode",
                    f"-env:UserInstallation={profile.resolve().as_uri()}",
                    "--convert-to",
                    "pdf",
                    "--outdir",
                    directory,
                    str(source),
                ],
                shell=False,
                capture_output=True,
                timeout=180,
                check=False,
            )
        except subprocess.TimeoutExpired as exc:
            raise AerError(
                "COMMAND_TIMEOUT",
                "LibreOffice conversion timed out.",
                "artifact.convert",
                str(source),
            ) from exc
        generated = Path(directory) / f"{source.stem}.pdf"
        if completed.returncode != 0 or not generated.is_file():
            raise AerError(
                "COMMAND_FAILED",
                "LibreOffice conversion failed.",
                "artifact.convert",
                str(source),
                {"exit_code": completed.returncode},
            )
        _copy_bounded(generated, output)


def _csv_to_xlsx(source: Path, output: Path) -> None:
    if source.stat().st_size > MAX_TABULAR_FILE_BYTES:
        raise AerError(
            "LIMIT_EXCEEDED",
            "Tabular input exceeds the conversion size limit.",
            "artifact.convert",
            str(source),
            {"bytes": source.stat().st_size, "limit": MAX_TABULAR_FILE_BYTES},
        )
    delimiter = "\t" if source.suffix.lower() == ".tsv" else ","
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    materialized_cells = 0
    with source.open("r", encoding="utf-8-sig", newline="") as handle:
        for row_number, row in enumerate(csv.reader(handle, delimiter=delimiter), start=1):
            materialized_cells += len(row)
            if row_number > MAX_TABULAR_ROWS or materialized_cells > MAX_TABULAR_CELLS:
                raise AerError(
                    "LIMIT_EXCEEDED",
                    "Tabular input exceeds the row or cell conversion limit.",
                    "artifact.convert",
                    str(source),
                    {
                        "rows": row_number,
                        "columns": len(row),
                        "cells": materialized_cells,
                        "row_limit": MAX_TABULAR_ROWS,
                        "cell_limit": MAX_TABULAR_CELLS,
                    },
                )
            sheet.append(row)
            for cell, value in zip(sheet[sheet.max_row], row, strict=True):
                if value.startswith(("=", "+", "-", "@")):
                    cell.data_type = "s"
                    cell.quotePrefix = True
    if sheet.max_row:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
    with tempfile.TemporaryDirectory(prefix="aer-xlsx-") as directory:
        temporary = Path(directory) / "converted.xlsx"
        workbook.save(temporary)
        _copy_bounded(temporary, output)


def _xlsx_to_csv(source: Path, output: Path) -> None:
    enforce_zip_expansion_limits(source, operation="artifact.convert", target=str(source))
    workbook = load_workbook(source, read_only=True, data_only=False)
    sheet = workbook.active
    total = 0
    delimiter = "\t" if output.suffix.casefold() == ".tsv" else ","
    with atomic_binary_writer(output) as destination:
        for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if row_number > MAX_TABULAR_ROWS or row_number * max(1, len(row)) > MAX_TABULAR_CELLS:
                raise AerError(
                    "LIMIT_EXCEEDED",
                    "Workbook exceeds the row or cell conversion limit.",
                    "artifact.convert",
                    str(source),
                )
            stream = io.StringIO(newline="")
            csv.writer(stream, delimiter=delimiter).writerow(
                ["" if value is None else value for value in row]
            )
            encoded = stream.getvalue().encode("utf-8")
            total += len(encoded)
            if total > MAX_TABULAR_FILE_BYTES:
                raise AerError(
                    "LIMIT_EXCEEDED",
                    "Converted tabular output exceeds the size limit.",
                    "artifact.convert",
                    str(output),
                )
            destination.write(encoded)


def _image_convert(source: Path, output: Path) -> None:
    format_by_suffix = {
        ".png": "PNG",
        ".jpg": "JPEG",
        ".jpeg": "JPEG",
        ".webp": "WEBP",
        ".tif": "TIFF",
        ".tiff": "TIFF",
    }
    format_name = format_by_suffix.get(output.suffix.lower())
    if not format_name:
        raise AerError(
            "UNSUPPORTED_FORMAT",
            "Unsupported image output format.",
            "artifact.convert",
            str(output),
        )
    try:
        with Image.open(source) as opened:
            if opened.width * opened.height > MAX_IMAGE_PIXELS:
                raise AerError(
                    "LIMIT_EXCEEDED",
                    "Image pixel count exceeds the safety limit.",
                    "artifact.convert",
                    str(source),
                )
            image = ImageOps.exif_transpose(opened)
            image.load()
    except AerError:
        raise
    except Exception as exc:
        raise AerError(
            "CORRUPT_FILE", f"Cannot open image: {exc}", "artifact.convert", str(source)
        ) from exc
    if format_name == "JPEG" and image.mode not in {"RGB", "L"}:
        background = Image.new("RGB", image.size, "white")
        background.paste(image, mask=image.getchannel("A") if image.mode == "RGBA" else None)
        image = background
    with atomic_binary_writer(output) as destination:
        image.save(destination, format=format_name)


def _pandoc(source: Path, output: Path) -> None:
    executable = shutil.which("pandoc")
    if not executable:
        raise _dependency("pandoc", "markup.convert", source)
    if source.stat().st_size > MAX_TEXT_FILE_BYTES:
        raise AerError(
            "LIMIT_EXCEEDED",
            "Markup source exceeds the conversion size limit.",
            "artifact.convert",
            str(source),
        )
    try:
        markup = source.read_text(encoding="utf-8")
    except (OSError, UnicodeError) as exc:
        raise AerError(
            "CORRUPT_FILE", "Markup source must be readable UTF-8.", "artifact.convert", str(source)
        ) from exc
    if re.search(r"(?i)(?:\b[a-z][a-z0-9+.-]*:)?//|\bfile:", markup):
        raise AerError(
            "UNSUPPORTED_FORMAT",
            "Remote and file URL references are not fetched during markup conversion.",
            "artifact.convert",
            str(source),
            suggested_action="Replace external references with local, explicitly packaged assets.",
        )
    with tempfile.TemporaryDirectory(prefix="aer-pandoc-") as directory:
        temporary = Path(directory) / output.name
        try:
            completed = subprocess.run(
                [executable, str(source), "-o", str(temporary)],
                shell=False,
                capture_output=True,
                timeout=120,
                check=False,
            )
        except subprocess.TimeoutExpired as exc:
            raise AerError(
                "COMMAND_TIMEOUT", "Pandoc conversion timed out.", "artifact.convert", str(source)
            ) from exc
        if completed.returncode != 0 or not temporary.is_file():
            raise AerError(
                "COMMAND_FAILED",
                "Pandoc conversion failed.",
                "artifact.convert",
                str(source),
                {"exit_code": completed.returncode},
            )
        _copy_bounded(temporary, output)


def convert_file(source: Path, output: Path) -> dict[str, Any]:
    source = ensure_regular_input(source, operation="artifact.convert")
    output = prepare_output_path(output, operation="artifact.convert")
    if source.resolve() == output:
        raise AerError(
            "CONFLICT", "Conversion output must differ from input.", "artifact.convert", str(output)
        )
    source_suffix, output_suffix = source.suffix.lower(), output.suffix.lower()
    if source_suffix in {".docx", ".pptx", ".xlsx"} and output_suffix == ".pdf":
        _office_to_pdf(source, output)
    elif source_suffix in {".csv", ".tsv"} and output_suffix == ".xlsx":
        _csv_to_xlsx(source, output)
    elif source_suffix == ".xlsx" and output_suffix in {".csv", ".tsv"}:
        _xlsx_to_csv(source, output)
    elif source_suffix in {".png", ".jpg", ".jpeg", ".webp", ".tif", ".tiff"} and output_suffix in {
        ".png",
        ".jpg",
        ".jpeg",
        ".webp",
        ".tif",
        ".tiff",
    }:
        _image_convert(source, output)
    elif source_suffix in {".md", ".markdown", ".html", ".htm"}:
        _pandoc(source, output)
    else:
        raise AerError(
            "UNSUPPORTED_FORMAT",
            "This conversion pair is not supported.",
            "artifact.convert",
            str(source),
            {"source_extension": source_suffix, "output_extension": output_suffix},
        )
    return {
        "input": str(source),
        "output": str(output),
        "bytes": output.stat().st_size,
        "sha256": sha256_file(output),
    }
