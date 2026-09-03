from __future__ import annotations

import io
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference, ScatterChart, Series
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import coordinate_to_tuple, get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.table import Table, TableStyleInfo

from aer.artifacts.workbook.selectors import stable_cell_name, stable_sheet_name
from aer.errors import AerError
from aer.limits import (
    MAX_ARTIFACT_ELEMENTS,
    MAX_TABULAR_CELLS,
    MAX_TABULAR_ROWS,
    MAX_WORKBOOK_SHEETS,
)
from aer.paths import atomic_write_bytes


def _validate_workbook_limits(sheets: list[Any]) -> None:
    if len(sheets) > MAX_WORKBOOK_SHEETS:
        raise AerError(
            "LIMIT_EXCEEDED",
            "Workbook exceeds the sheet-count limit.",
            "workbook.build",
            "/sheets",
            {"sheets": len(sheets), "limit": MAX_WORKBOOK_SHEETS},
        )
    total_cells = 0
    total_auxiliary = 0
    for index, sheet_spec in enumerate(sheets):
        if not isinstance(sheet_spec, dict):
            continue
        rows = sheet_spec.get("rows", [])
        columns = sheet_spec.get("columns")
        cells = sheet_spec.get("cells", [])
        charts = sheet_spec.get("charts", [])
        conditionals = sheet_spec.get("conditional_formats", [])
        if (
            not isinstance(rows, list)
            or not isinstance(cells, list)
            or (columns is not None and not isinstance(columns, list))
        ):
            raise AerError(
                "INVALID_SPEC",
                "Workbook columns, rows, and cells must be arrays.",
                "workbook.build",
                f"/sheets/{index}",
            )
        if not isinstance(charts, list) or not isinstance(conditionals, list):
            raise AerError(
                "INVALID_SPEC",
                "Workbook charts and conditional formats must be arrays.",
                "workbook.build",
                f"/sheets/{index}",
            )
        header_width = len(columns) if columns is not None else 0
        row_width = 0
        for row_index, row in enumerate(rows):
            if isinstance(row, dict):
                width = header_width or len(row)
            elif isinstance(row, list):
                width = len(row)
            else:
                raise AerError(
                    "INVALID_SPEC",
                    "Workbook rows must be arrays or objects.",
                    "workbook.build",
                    f"/sheets/{index}/rows/{row_index}",
                )
            row_width = max(row_width, width)
        header_rows = 1 if header_width or (rows and isinstance(rows[0], dict)) else 0
        maximum_row = len(rows) + header_rows
        maximum_column = max(header_width, row_width)
        for cell_index, cell_spec in enumerate(cells):
            if not isinstance(cell_spec, dict):
                raise AerError(
                    "INVALID_SPEC",
                    "Workbook cell entries must be objects.",
                    "workbook.build",
                    f"/sheets/{index}/cells/{cell_index}",
                )
            if "address" not in cell_spec:
                raise AerError(
                    "INVALID_SPEC",
                    "Workbook cell entries require an address.",
                    "workbook.build",
                    "/address",
                )
            try:
                cell_row, cell_column = coordinate_to_tuple(str(cell_spec["address"]))
            except ValueError as exc:
                raise AerError(
                    "INVALID_SPEC",
                    "Workbook cell address is invalid.",
                    "workbook.build",
                    f"/sheets/{index}/cells/{cell_index}/address",
                ) from exc
            maximum_row = max(maximum_row, cell_row)
            maximum_column = max(maximum_column, cell_column)
        if maximum_row > MAX_TABULAR_ROWS:
            raise AerError(
                "LIMIT_EXCEEDED",
                "Workbook sheet exceeds the row limit.",
                "workbook.build",
                f"/sheets/{index}",
                {"rows": maximum_row, "limit": MAX_TABULAR_ROWS},
            )
        total_cells += maximum_row * maximum_column
        total_auxiliary += len(charts) + len(conditionals)
        if total_cells > MAX_TABULAR_CELLS:
            raise AerError(
                "LIMIT_EXCEEDED",
                "Workbook exceeds the materialized-cell limit.",
                "workbook.build",
                f"/sheets/{index}",
                {"cells": total_cells, "limit": MAX_TABULAR_CELLS},
            )
        if total_auxiliary > MAX_ARTIFACT_ELEMENTS:
            raise AerError(
                "LIMIT_EXCEEDED",
                "Workbook exceeds the chart and conditional-format count limit.",
                "workbook.build",
                f"/sheets/{index}",
                {"items": total_auxiliary, "limit": MAX_ARTIFACT_ELEMENTS},
            )


def _chart(sheet: Any, chart_spec: dict[str, Any]) -> None:
    chart_type = str(chart_spec.get("type", "bar"))
    if chart_type == "line":
        chart: Any = LineChart()
    elif chart_type == "pie":
        chart = PieChart()
    elif chart_type == "scatter":
        chart = ScatterChart()
    else:
        chart = BarChart()
        if chart_type == "horizontal-bar":
            chart.type = "bar"
    chart.title = str(chart_spec.get("title", ""))
    minimum_row = int(chart_spec.get("min_row", 1))
    maximum_row = int(chart_spec.get("max_row", sheet.max_row))
    minimum_column = int(chart_spec.get("min_col", 2))
    maximum_column = int(chart_spec.get("max_col", sheet.max_column))
    categories_column = int(chart_spec.get("categories_col", 1))
    if chart_type == "scatter":
        x_values = Reference(
            sheet, min_col=categories_column, min_row=minimum_row + 1, max_row=maximum_row
        )
        for column in range(minimum_column, maximum_column + 1):
            values = Reference(sheet, min_col=column, min_row=minimum_row, max_row=maximum_row)
            chart.series.append(Series(values, x_values, title_from_data=True))
    else:
        data = Reference(
            sheet,
            min_col=minimum_column,
            max_col=maximum_column,
            min_row=minimum_row,
            max_row=maximum_row,
        )
        categories = Reference(
            sheet, min_col=categories_column, min_row=minimum_row + 1, max_row=maximum_row
        )
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
    sheet.add_chart(chart, str(chart_spec.get("anchor", "H2")))


def build_workbook(
    spec: dict[str, Any], output: Path, *, spec_dir: Path
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    del spec_dir
    sheets = spec.get("sheets", spec.get("content"))
    if not isinstance(sheets, list) or not sheets:
        raise AerError(
            "INVALID_SPEC",
            "Workbook sheets must be a non-empty array.",
            "workbook.build",
            "/sheets",
        )
    _validate_workbook_limits(sheets)
    workbook = Workbook()
    workbook.remove(workbook.active)
    elements: list[dict[str, Any]] = []
    seen_titles: set[str] = set()
    seen_sheet_ids: set[str] = set()
    seen_defined_names: set[str] = set()
    for index, sheet_spec in enumerate(sheets):
        if not isinstance(sheet_spec, dict):
            raise AerError(
                "INVALID_SPEC",
                "Each sheet must be an object.",
                "workbook.build",
                f"/sheets/{index}",
            )
        title = str(sheet_spec.get("name", sheet_spec.get("id", f"Sheet{index + 1}")))
        stable_id = str(sheet_spec.get("id", title))
        if title in seen_titles:
            raise AerError(
                "INVALID_SPEC",
                "Sheet names must be unique.",
                "workbook.build",
                f"/sheets/{index}/name",
            )
        sheet_defined_name = stable_sheet_name(stable_id)
        if (
            not stable_id
            or "/" in stable_id
            or stable_id in seen_sheet_ids
            or sheet_defined_name.casefold() in seen_defined_names
        ):
            raise AerError(
                "INVALID_SPEC",
                "Sheet IDs must be unique after stable-selector normalization and cannot contain '/'.",
                "workbook.build",
                f"/sheets/{index}/id",
            )
        seen_titles.add(title)
        seen_sheet_ids.add(stable_id)
        seen_defined_names.add(sheet_defined_name.casefold())
        sheet = workbook.create_sheet(title)
        workbook.defined_names.add(DefinedName(sheet_defined_name, attr_text=f"'{title}'!$A$1"))
        rows = sheet_spec.get("rows", [])
        columns = sheet_spec.get("columns")
        if rows and isinstance(rows[0], dict):
            headers = [str(value) for value in (columns or list(rows[0].keys()))]
            sheet.append(headers)
            for row in rows:
                sheet.append([row.get(header) for header in headers])
        else:
            if columns:
                sheet.append([str(value) for value in columns])
            for row in rows:
                sheet.append(list(row))
        tabular_max_row = sheet.max_row
        tabular_max_column = sheet.max_column
        cell_elements: list[dict[str, Any]] = []
        seen_cell_ids: set[str] = set()
        for cell_index, cell_spec in enumerate(sheet_spec.get("cells", [])):
            address = str(cell_spec["address"])
            cell = sheet[address]
            cell.value = cell_spec.get("formula", cell_spec.get("value"))
            if cell_spec.get("number_format"):
                cell.number_format = str(cell_spec["number_format"])
            raw_cell_id = cell_spec.get("id")
            if raw_cell_id is not None:
                cell_id = str(raw_cell_id)
                name = stable_cell_name(stable_id, cell_id)
                if (
                    not cell_id
                    or "/" in cell_id
                    or cell_id in seen_cell_ids
                    or name.casefold() in seen_defined_names
                ):
                    raise AerError(
                        "INVALID_SPEC",
                        "Cell IDs must be unique after stable-selector normalization and cannot contain '/'.",
                        "workbook.build",
                        f"/sheets/{index}/cells/{cell_index}/id",
                    )
                seen_cell_ids.add(cell_id)
                seen_defined_names.add(name.casefold())
                workbook.defined_names.add(
                    DefinedName(name, attr_text=f"'{title}'!{cell.coordinate}")
                )
                cell_elements.append(
                    {
                        "id": cell_id,
                        "type": "cell",
                        "address": cell.coordinate,
                        "selector": f"sheet:id={stable_id}/cell:id={cell_id}",
                    }
                )
        if sheet.max_row > 0 and sheet.max_column > 0:
            for cell in sheet[1]:
                cell.font = Font(name="Noto Sans CJK KR", bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="193A5A")
                cell.alignment = Alignment(horizontal="center")
            for row in sheet.iter_rows(min_row=2):
                for cell in row:
                    cell.font = Font(name="Noto Sans CJK KR", size=10)
        frozen = sheet_spec.get("freeze", sheet_spec.get("frozen_pane"))
        if frozen:
            sheet.freeze_panes = str(frozen)
        if sheet_spec.get("auto_filter", False) and sheet.max_row:
            sheet.auto_filter.ref = sheet.dimensions
        widths = sheet_spec.get("column_widths", {})
        for key, value in widths.items():
            sheet.column_dimensions[str(key)].width = float(value)
        if sheet_spec.get("auto_width", True):
            for column in range(1, sheet.max_column + 1):
                maximum = max(
                    (
                        len(str(sheet.cell(row, column).value or ""))
                        for row in range(1, min(sheet.max_row, 200) + 1)
                    ),
                    default=8,
                )
                sheet.column_dimensions[get_column_letter(column)].width = min(
                    max(maximum + 2, 8), 48
                )
        if sheet_spec.get("table") and tabular_max_row >= 2:
            table_config = sheet_spec["table"]
            table_name = (
                str(table_config.get("name", f"Table{index + 1}"))
                if isinstance(table_config, dict)
                else f"Table{index + 1}"
            )
            table_ref = f"A1:{get_column_letter(tabular_max_column)}{tabular_max_row}"
            table = Table(displayName=table_name, ref=table_ref)
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False
            )
            sheet.add_table(table)
        for conditional in sheet_spec.get("conditional_formats", []):
            target = str(conditional["range"])
            kind = str(conditional.get("type", "color-scale"))
            if kind == "color-scale":
                sheet.conditional_formatting.add(
                    target,
                    ColorScaleRule(
                        start_type="min",
                        start_color="F8696B",
                        mid_type="percentile",
                        mid_value=50,
                        mid_color="FFEB84",
                        end_type="max",
                        end_color="63BE7B",
                    ),
                )
            elif kind == "cell":
                sheet.conditional_formatting.add(
                    target,
                    CellIsRule(
                        operator=str(conditional.get("operator", "greaterThan")),
                        formula=[str(conditional.get("value", 0))],
                        fill=PatternFill("solid", fgColor=str(conditional.get("color", "FFEB84"))),
                    ),
                )
        for chart_spec in sheet_spec.get("charts", []):
            _chart(sheet, chart_spec)
        if sheet_spec.get("named_range"):
            name = str(sheet_spec["named_range"])
            if name.casefold() in seen_defined_names:
                raise AerError(
                    "INVALID_SPEC",
                    "Named ranges cannot replace an AER stable selector.",
                    "workbook.build",
                    f"/sheets/{index}/named_range",
                )
            seen_defined_names.add(name.casefold())
            workbook.defined_names.add(DefinedName(name, attr_text=f"'{title}'!{sheet.dimensions}"))
        sheet_element: dict[str, Any] = {
            "id": stable_id,
            "type": "sheet",
            "name": title,
            "selector": f"sheet:id={stable_id}",
        }
        if cell_elements:
            sheet_element["children"] = cell_elements
        elements.append(sheet_element)
    workbook.calculation.fullCalcOnLoad = True
    buffer = io.BytesIO()
    workbook.save(buffer)
    atomic_write_bytes(output, buffer.getvalue())
    return elements, [
        {
            "code": "FORMULAS_NOT_CALCULATED",
            "message": "Formula strings are preserved; AER does not calculate workbook formulas.",
        }
    ]
